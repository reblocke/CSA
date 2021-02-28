from ReadExcel import *
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib.sankey as sankey
import matplotlib.gridspec as gridspec
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.preprocessing import LabelEncoder


def summary_stats(df):
    print("\nAge Summary Statistics:\n")
    print(str(df['Age'].describe()))

    print("\nSex Counts:\n")
    print(str(df['Sex'].value_counts()))

    print("\nRace Counts:\n")
    print(str(df['Race'].value_counts()))

    print("\nSmoking Counts:\n")
    print(str(df['Smoking'].value_counts()))

    print("\nBMI Summary Statistics:\n")
    print(str(df['BMI'].describe()))

    print("\nComorbidity Combination Counts:\n")
    print(str(df['Comorb'].value_counts()))
    print("\nComorbidity Counts:\n")
    print(str(histo_comorbs_includes(df)))

    print("\nHeart Combination Counts:\n")
    print(str(df['Heart'].value_counts()))
    print("\nHeart Counts:\n")
    print(str(histo_heart_includes(df)))

    print("\nCNS Combination Counts:\n")
    print(str(df['CNS'].value_counts()))
    print("\nCNS Counts:\n")
    print(str(histo_cns_includes(df)))

    print("\nAHI Summary Statistics:\n")
    print(str(df['AHI'].describe()))

    print("\nBase Dx Counts:\n")
    print(str(df['BaseDx'].value_counts()))

    print("\nEtiology Combination Counts:\n")
    print(str(df['PostDx'].value_counts()))
    print("\nEtiology Counts:\n")
    print(str(histo_dx_includes(df)))

    print("\nFinal Treatment Counts:\n")
    print(str(df['FinalTx'].value_counts()))

    print("\nOutcome Counts:\n")
    print(str(df['Outcome'].value_counts()))


def new_make_tables(df):
    """makes 3 tables on:
    Demographics for each
    Outcome for each
    Etiology for each

    Each = stratified by <10%, 10-50%, 50-90%, 90+%

    Formatted to be more readable than older version
    """

    CSA_pure = df.loc[df['BaseDx'] == "Pure CSA"]
    CSA_predom = df.loc[df['BaseDx'] == "Predominantly CSA"]
    OSA_predom = df.loc[df['BaseDx'] == "Combined OSA/CSA"]
    OSA_pure = df.loc[df['BaseDx'] == "Mainly OSA"]

    num_total = len(df.index)
    num_csa_pure = len(CSA_pure.index)
    num_csa_predom = len(CSA_predom.index)
    num_osa_predom = len(OSA_predom.index)
    num_osa_pure = len(OSA_pure.index)

    column_labels = ['All',
                     'Pure CSA (90+% CSA)',
                     'Predominantly CSA (50-90% CSA)',
                     'Combined OSA/CSA (10-49.9% CSA)',
                     'Mainly OSA (<10% CSA)']

    workbook = Workbook()

    # Table 1: Demographics - each list is row
    demo_row_labels = []
    demographics = []

    # Age
    demo_row_labels.append('AGE')
    demographics.append((std_string(df['Age'].describe()), std_string(CSA_pure['Age'].describe()),
                     std_string(CSA_predom['Age'].describe()), std_string(OSA_predom['Age'].describe()),
                     std_string(OSA_pure['Age'].describe())))

    demo_row_labels.append("")
    demographics.append(("", "", "", "", "")) # empty row

    # Sex
    demo_row_labels.append('GENDER')
    demographics.append(("", "", "", "", "")) # empty row

    for sex in df['Sex'].value_counts().keys():
        demo_row_labels.append(sex)
        demographics.append((count_string_indiv(df['Sex'].value_counts()[sex], num_total),
                            count_string_indiv(CSA_pure['Sex'].value_counts(dropna=False)[sex], num_csa_pure),
                            count_string_indiv(CSA_predom['Sex'].value_counts(dropna=False)[sex], num_csa_predom),
                            count_string_indiv(OSA_predom['Sex'].value_counts(dropna=False)[sex], num_osa_predom),
                            count_string_indiv(OSA_pure['Sex'].value_counts(dropna=False)[sex], num_osa_pure)))

    demo_row_labels.append("")
    demographics.append(("", "", "", "", "")) # empty row

    # Race
    demo_row_labels.append('RACE/ETHNICITY')
    demographics.append(("", "", "", "", "")) # empty row

    for race in df['Race'].value_counts().keys():
        demo_row_labels.append(race)
        count_total = df['Race'].value_counts()[race]

        try:
            count_csa_pure = CSA_pure['Race'].value_counts()[race]
        except(KeyError):
            count_csa_pure = 0
        try:
            count_csa_predom = CSA_predom['Race'].value_counts()[race]
        except(KeyError):
            count_csa_predom = 0
        try:
            count_osa_predom = OSA_predom['Race'].value_counts()[race]
        except(KeyError):
            count_osa_predom = 0
        try:
            count_osa_pure = OSA_pure['Race'].value_counts(dropna=False)[race]
        except(KeyError):
            count_osa_pure = 0

        demographics.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    demo_row_labels.append("")
    demographics.append(("", "", "", "", "")) # empty row

    # Smoking
    demo_row_labels.append('SMOKING')
    demographics.append(("", "", "", "", "")) # empty row

    for status in df['Smoking'].value_counts().keys():
        demo_row_labels.append(status)
        count_total = df['Smoking'].value_counts()[status]

        try:
            count_csa_pure = CSA_pure['Smoking'].value_counts()[status]
        except(KeyError):
            count_csa_pure = 0
        try:
            count_csa_predom = CSA_predom['Smoking'].value_counts()[status]
        except(KeyError):
            count_csa_predom = 0
        try:
            count_osa_predom = OSA_predom['Smoking'].value_counts()[status]
        except(KeyError):
            count_osa_predom = 0
        try:
            count_osa_pure = OSA_pure['Smoking'].value_counts(dropna=False)[status]
        except(KeyError):
            count_osa_pure = 0

        demographics.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    demo_row_labels.append("")
    demographics.append(("", "", "", "", ""))  # empty row

    # BMI
    demo_row_labels.append('BMI')
    # demographics.append((iqr_string(df['BMI'].describe()), iqr_string(CSA_pure['BMI'].describe()),
    #                 iqr_string(CSA_predom['BMI'].describe()), iqr_string(OSA_predom['BMI'].describe()),
    #                 iqr_string(OSA_pure['BMI'].describe())))

    demographics.append((std_string(df['BMI'].describe()), std_string(CSA_pure['BMI'].describe()),
                     std_string(CSA_predom['BMI'].describe()), std_string(OSA_predom['BMI'].describe()),
                     std_string(OSA_pure['BMI'].describe())))


    demo_row_labels.append("")
    demographics.append(("", "", "", "", "")) # empty row

    # Comorbidities
    demo_row_labels.append('COMORBIDITIES: [ ] manually add CAD to this in place of None')
    demographics.append(("", "", "", "", "")) # empty row

    for status in histo_comorbs_includes(df).keys():
        demo_row_labels.append(status)
        count_total = histo_comorbs_includes(df)[status] # Automatically returns all keys (including a value if 0)
        count_csa_pure = histo_comorbs_includes(CSA_pure)[status]
        count_csa_predom = histo_comorbs_includes(CSA_predom)[status]
        count_osa_predom = histo_comorbs_includes(OSA_predom)[status]
        count_osa_pure = histo_comorbs_includes(OSA_pure)[status]

        demographics.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    demo_row_labels.append("")
    demographics.append(("", "", "", "", ""))  # empty row

    demo_row_labels.append('OSA SEVERITY BY AHI')
    demographics.append(("", "", "", "", "")) # empty row

    # AHI
    demo_row_labels.append('AHI')
    # demographics.append((iqr_string(df['AHI'].describe()), iqr_string(CSA_pure['AHI'].describe()),
    #                  iqr_string(CSA_predom['AHI'].describe()), iqr_string(OSA_predom['AHI'].describe()),
    #                  iqr_string(OSA_pure['AHI'].describe())))

    demographics.append((std_string(df['AHI'].describe()), std_string(CSA_pure['AHI'].describe()),
                     std_string(CSA_predom['AHI'].describe()), std_string(OSA_predom['AHI'].describe()),
                     std_string(OSA_pure['AHI'].describe())))

    for severity in df['AHI_label'].value_counts().keys():
        demo_row_labels.append(severity)
        count_total = df['AHI_label'].value_counts()[severity]

        try:
            count_csa_pure = CSA_pure['AHI_label'].value_counts()[severity]
        except(KeyError):
            count_csa_pure = 0
        try:
            count_csa_predom = CSA_predom['AHI_label'].value_counts()[severity]
        except(KeyError):
            count_csa_predom = 0
        try:
            count_osa_predom = OSA_predom['AHI_label'].value_counts()[severity]
        except(KeyError):
            count_osa_predom = 0
        try:
            count_osa_pure = OSA_pure['AHI_label'].value_counts(dropna=False)[severity]
        except(KeyError):
            count_osa_pure = 0

        demographics.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    demo_row_labels.append("")
    demographics.append(("", "", "", "", ""))  # empty row

    # Diagnostic Test

    demo_row_labels.append('DIAGNOSTIC TEST')
    demographics.append(("", "", "", "", "")) # empty row

    for status in df['StudyType'].value_counts().keys():
        demo_row_labels.append(status)
        count_total = df['StudyType'].value_counts()[status]

        try:
            count_csa_pure = CSA_pure["StudyType"].value_counts()[status]
        except(KeyError):
            count_csa_pure = 0
        try:
            count_csa_predom = CSA_predom["StudyType"].value_counts()[status]
        except(KeyError):
            count_csa_predom = 0
        try:
            count_osa_predom = OSA_predom["StudyType"].value_counts()[status]
        except(KeyError):
            count_osa_predom = 0
        try:
            count_osa_pure = OSA_pure["StudyType"].value_counts(dropna=False)[status]
        except(KeyError):
            count_osa_pure = 0

        demographics.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    demo_row_labels.append("")
    demographics.append(("", "", "", "", ""))  # empty row
    demo_row_labels.append("")
    demographics.append(("*Note:", "CKD defined as baseline serum creatinine of 2 mg/dl or use of renal replacement therapy", "", "", ""))

    demographics_df = pd.DataFrame(demographics, columns=column_labels, index=demo_row_labels)

    demographic_worksheet = workbook.worksheets[0]
    demographic_worksheet.title = "Demographics"

    for r in dataframe_to_rows(demographics_df, index=True, header=True):
        demographic_worksheet.append(r)

    for cell in demographic_worksheet['A'] + demographic_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')

    # Table 2 - Etiology
    etio_row_labels = []
    etiologies = []

    # Category:
    etio_row_labels.append("CAUSE OF CENTRAL SLEEP APNEAS")
    etiologies.append(("", "", "", "", ""))  # empty row

    for key in histo_dx_includes(df).keys():
        etio_row_labels.append(key)
        count_total = histo_dx_includes(df)[key] # Automatically returns all keys (including a value if 0)
        count_csa_pure = histo_dx_includes(CSA_pure)[key]
        count_csa_predom = histo_dx_includes(CSA_predom)[key]
        count_osa_predom = histo_dx_includes(OSA_predom)[key]
        count_osa_pure = histo_dx_includes(OSA_pure)[key]

        etiologies.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    etio_row_labels.append("")
    etiologies.append(("", "", "", "", ""))  # empty row

    # CV comorbs:
    # TODO: the CV comorbidity counts should be in table 1, not 2 (not taken as etiology) - currently changed in post
    etio_row_labels.append("CARDIOVASCULAR CAUSES OF CSA")
    etiologies.append(("", "", "", "", ""))  # empty row

    for key in histo_heart_includes(df).keys():
        etio_row_labels.append(key)
        count_total = histo_heart_includes(df)[key] # Automatically returns all keys (including a value if 0)
        count_csa_pure = histo_heart_includes(CSA_pure)[key]
        count_csa_predom = histo_heart_includes(CSA_predom)[key]
        count_osa_predom = histo_heart_includes(OSA_predom)[key]
        count_osa_pure = histo_heart_includes(OSA_pure)[key]

        etiologies.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    etio_row_labels.append("")
    etiologies.append(("", "", "", "", ""))  # empty row

    # CNS comorbs:
    etio_row_labels.append("CNS CAUSES OF CSA")
    etiologies.append(("", "", "", "", ""))  # empty row

    for key in histo_cns_includes(df).keys():
        etio_row_labels.append(key)
        count_total = histo_cns_includes(df)[key] # Automatically returns all keys (including a value if 0)
        count_csa_pure = histo_cns_includes(CSA_pure)[key]
        count_csa_predom = histo_cns_includes(CSA_predom)[key]
        count_osa_predom = histo_cns_includes(OSA_predom)[key]
        count_osa_pure = histo_cns_includes(OSA_pure)[key]

        etiologies.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    etio_row_labels.append("")
    etiologies.append(("", "", "", "", ""))  # empty row
    etio_row_labels.append("")
    etiologies.append(("*Note:", "CAD does not count as causative of CSA and is listed in table 1", "", "", ""))
    etio_row_labels.append("")
    etiologies.append(("*Note2:", "does not sum to total- If multiple comorbidities present, counted toward each", "", "", ""))

    etiology_df = pd.DataFrame(etiologies, columns=column_labels, index=etio_row_labels)

    etiology_worksheet = workbook.create_sheet(title="Etiology", index=1)

    for r in dataframe_to_rows(etiology_df, index=True, header=True):
        etiology_worksheet.append(r)

    for cell in etiology_worksheet['A'] + etiology_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center')

    # Table 3 - Outcome
    outcome_row_labels = []
    outcome = []

    # 'Initial Treatment'
    outcome_row_labels.append('INITIAL TREATMENT')
    outcome.append(("", "", "", "", "")) # empty row

    for key in df['InitTx'].value_counts().keys():
        outcome_row_labels.append(key)
        count_total = df['InitTx'].value_counts()[key]

        try:
            count_csa_pure = CSA_pure['InitTx'].value_counts()[key]
        except(KeyError):
            count_csa_pure = 0
        try:
            count_csa_predom = CSA_predom['InitTx'].value_counts()[key]
        except(KeyError):
            count_csa_predom = 0
        try:
            count_osa_predom = OSA_predom['InitTx'].value_counts()[key]
        except(KeyError):
            count_osa_predom = 0
        try:
            count_osa_pure = OSA_pure['InitTx'].value_counts()[key]
        except(KeyError):
            count_osa_pure = 0

        outcome.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    outcome_row_labels.append("")
    outcome.append(("", "", "", "", ""))  # empty row

    # 'Final Treatment',
    outcome_row_labels.append('FINAL TREATMENT')
    outcome.append(("", "", "", "", "")) # empty row

    for key in df['FinalTx'].value_counts().keys():
        outcome_row_labels.append(key)
        count_total = df['FinalTx'].value_counts()[key]

        try:
            count_csa_pure = CSA_pure['FinalTx'].value_counts()[key]
        except(KeyError):
            count_csa_pure = 0
        try:
            count_csa_predom = CSA_predom['FinalTx'].value_counts()[key]
        except(KeyError):
            count_csa_predom = 0
        try:
            count_osa_predom = OSA_predom['FinalTx'].value_counts()[key]
        except(KeyError):
            count_osa_predom = 0
        try:
            count_osa_pure = OSA_pure['FinalTx'].value_counts()[key]
        except(KeyError):
            count_osa_pure = 0

        outcome.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    outcome_row_labels.append("")
    outcome.append(("", "", "", "", ""))  # empty row


    # 'Outcome'
    outcome_row_labels.append('OUTCOME OF INITIAL PAP THERAPY')
    outcome.append(("", "", "", "", "")) # empty row

    for key in df['Outcome'].value_counts().keys():
        outcome_row_labels.append(key)
        count_total = df['Outcome'].value_counts()[key]

        try:
            count_csa_pure = CSA_pure['Outcome'].value_counts()[key]
        except(KeyError):
            count_csa_pure = 0
        try:
            count_csa_predom = CSA_predom['Outcome'].value_counts()[key]
        except(KeyError):
            count_csa_predom = 0
        try:
            count_osa_predom = OSA_predom['Outcome'].value_counts()[key]
        except(KeyError):
            count_osa_predom = 0
        try:
            count_osa_pure = OSA_pure['Outcome'].value_counts()[key]
        except(KeyError):
            count_osa_pure = 0

        outcome.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_csa_pure, num_csa_pure),
                            count_string_indiv(count_csa_predom, num_csa_predom),
                            count_string_indiv(count_osa_predom, num_osa_predom),
                            count_string_indiv(count_osa_pure, num_osa_pure)))

    outcome_df = pd.DataFrame(outcome, columns=column_labels, index=outcome_row_labels)

    outcome_worksheet = workbook.create_sheet(title="Outcome", index=2)

    for r in dataframe_to_rows(outcome_df, index=True, header=True):
        outcome_worksheet.append(r)

    for cell in outcome_worksheet['A'] + outcome_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center')

    # Table 4 - Outcome by etiology

    neurologic_df = df.loc[df['PostDx'].str.contains("Neurologic")].sort_values(by='Outcome')
    cardiac_df = df.loc[df['PostDx'].str.contains("Cardiac")].sort_values(by='Outcome')
    medication_df = df.loc[df['PostDx'].str.contains("Medication")].sort_values(by='Outcome')
    tecsa_df = df.loc[df['PostDx'].str.contains("TECSA")].sort_values(by='Outcome')
    # osacsa_df = df.loc[df['PostDx'].str.contains("OSA-CSA")].sort_values(by='Outcome')
    primary_df = df.loc[df['PostDx'].str.contains("Primary")].sort_values(by='Outcome')

    num_neurologic = len(neurologic_df.index)
    num_cardiac = len(cardiac_df.index)
    num_medication = len(medication_df.index)
    num_tecsa = len(tecsa_df.index)
    # num_osacsa = len(osacsa_df.index)
    num_primary = len(primary_df.index)

    column_etio_labels = ['All',
                     'Neurologic Contributor',
                     'Cardiac Contributor',
                     'Medication Contributor',
                     'Treatment Emergent',
                     # 'OSA-associated Centrals',
                     'Primary CSA']

    outcome_etio_row_labels = []
    outcome_etio = []

    # Diagnostic Test

    outcome_etio_row_labels.append('DIAGNOSTIC TEST')
    outcome_etio.append(("", "", "", "", "", ""))  # , "")) # empty row

    for key in df["StudyType"].value_counts().keys():
        outcome_etio_row_labels.append(key)
        count_total = df["StudyType"].value_counts()[key]

        try:
            count_neurologic = neurologic_df["StudyType"].value_counts()[key]
        except(KeyError):
            count_neurologic = 0
        try:
            count_cardiac = cardiac_df["StudyType"].value_counts()[key]
        except(KeyError):
            count_cardiac = 0
        try:
            count_medication = medication_df["StudyType"].value_counts()[key]
        except(KeyError):
            count_medication = 0
        try:
            count_tecsa = tecsa_df["StudyType"].value_counts()[key]
        except(KeyError):
            count_tecsa = 0
        # try:
        #     count_osacsa = osacsa_df["StudyType"].value_counts()[key]
        # except(KeyError):
        #     count_osacsa = 0
        try:
            count_primary = primary_df["StudyType"].value_counts()[key]
        except(KeyError):
            count_primary = 0

        outcome_etio.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_neurologic, num_neurologic),
                            count_string_indiv(count_cardiac, num_cardiac),
                            count_string_indiv(count_medication, num_medication),
                            count_string_indiv(count_tecsa, num_tecsa),
                            # count_string_indiv(count_osacsa, num_osacsa),
                            count_string_indiv(count_primary, num_primary)))

    outcome_etio_row_labels.append("")
    outcome_etio.append(("", "", "", "", "", ""))  # , ""))  # empty row

    # 'Initial Treatment'

    outcome_etio_row_labels.append('INITIAL TREATMENT')
    outcome_etio.append(("", "", "", "", "", ""))  # , "")) # empty row

    for key in df['InitTx'].value_counts().keys():
        outcome_etio_row_labels.append(key)
        count_total = df['InitTx'].value_counts()[key]

        try:
            count_neurologic = neurologic_df['InitTx'].value_counts()[key]
        except(KeyError):
            count_neurologic = 0
        try:
            count_cardiac = cardiac_df['InitTx'].value_counts()[key]
        except(KeyError):
            count_cardiac = 0
        try:
            count_medication = medication_df['InitTx'].value_counts()[key]
        except(KeyError):
            count_medication = 0
        try:
            count_tecsa = tecsa_df['InitTx'].value_counts()[key]
        except(KeyError):
            count_tecsa = 0
        # try:
        #     count_osacsa = osacsa_df['InitTx'].value_counts()[key]
        # except(KeyError):
        #     count_osacsa = 0
        try:
            count_primary = primary_df['InitTx'].value_counts()[key]
        except(KeyError):
            count_primary = 0

        outcome_etio.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_neurologic, num_neurologic),
                            count_string_indiv(count_cardiac, num_cardiac),
                            count_string_indiv(count_medication, num_medication),
                            count_string_indiv(count_tecsa, num_tecsa),
                            # count_string_indiv(count_osacsa, num_osacsa),
                            count_string_indiv(count_primary, num_primary)))

    outcome_etio_row_labels.append("")
    outcome_etio.append(("", "", "", "", "", ""))  # , ""))  # empty row

    # 'Final Treatment'

    outcome_etio_row_labels.append('FINAL TREATMENT')
    outcome_etio.append(("", "", "", "", "", ""))  # , "")) # empty row

    for key in df['FinalTx'].value_counts().keys():
        outcome_etio_row_labels.append(key)
        count_total = df['FinalTx'].value_counts()[key]

        try:
            count_neurologic = neurologic_df['FinalTx'].value_counts()[key]
        except(KeyError):
            count_neurologic = 0
        try:
            count_cardiac = cardiac_df['FinalTx'].value_counts()[key]
        except(KeyError):
            count_cardiac = 0
        try:
            count_medication = medication_df['FinalTx'].value_counts()[key]
        except(KeyError):
            count_medication = 0
        try:
            count_tecsa = tecsa_df['FinalTx'].value_counts()[key]
        except(KeyError):
            count_tecsa = 0
        # try:
        #     count_osacsa = osacsa_df['FinalTx'].value_counts()[key]
        # except(KeyError):
        #     count_osacsa = 0
        try:
            count_primary = primary_df['FinalTx'].value_counts()[key]
        except(KeyError):
            count_primary = 0

        outcome_etio.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_neurologic, num_neurologic),
                            count_string_indiv(count_cardiac, num_cardiac),
                            count_string_indiv(count_medication, num_medication),
                            count_string_indiv(count_tecsa, num_tecsa),
                            #  count_string_indiv(count_osacsa, num_osacsa),
                            count_string_indiv(count_primary, num_primary)))

    outcome_etio_row_labels.append("")
    outcome_etio.append(("", "", "", "", "", ""))  # , ""))  # empty row

    # 'Outcome'

    outcome_etio_row_labels.append('OUTCOME OF INITIAL PAP THERAPY')
    outcome_etio.append(("", "", "", "", "", ""))  # , "")) # empty row

    for key in df['Outcome'].value_counts().keys():
        outcome_etio_row_labels.append(key)
        count_total = df['Outcome'].value_counts()[key]

        try:
            count_neurologic = neurologic_df['Outcome'].value_counts()[key]
        except(KeyError):
            count_neurologic = 0
        try:
            count_cardiac = cardiac_df['Outcome'].value_counts()[key]
        except(KeyError):
            count_cardiac = 0
        try:
            count_medication = medication_df['Outcome'].value_counts()[key]
        except(KeyError):
            count_medication = 0
        try:
            count_tecsa = tecsa_df['Outcome'].value_counts()[key]
        except(KeyError):
            count_tecsa = 0
        # try:
        #     count_osacsa = osacsa_df['Outcome'].value_counts()[key]
        # except(KeyError):
        #     count_osacsa = 0
        try:
            count_primary = primary_df['Outcome'].value_counts()[key]
        except(KeyError):
            count_primary = 0

        outcome_etio.append((count_string_indiv(count_total, num_total),
                            count_string_indiv(count_neurologic, num_neurologic),
                            count_string_indiv(count_cardiac, num_cardiac),
                            count_string_indiv(count_medication, num_medication),
                            count_string_indiv(count_tecsa, num_tecsa),
                            # count_string_indiv(count_osacsa, num_osacsa),
                            count_string_indiv(count_primary, num_primary)))

    outcome_etio_df = pd.DataFrame(outcome_etio, columns=column_etio_labels, index=outcome_etio_row_labels)

    outcome_etio_worksheet = workbook.create_sheet(title="Outcome by Etiology", index=3)

    for r in dataframe_to_rows(outcome_etio_df, index=True, header=True):
        outcome_etio_worksheet.append(r)

    for cell in outcome_etio_worksheet['A'] + outcome_etio_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center')

    workbook.save("tables.xlsx")
    return


def makeTables(df):
    """makes 3 tables on:
    Demographics for each
    Outcome for each
    Etiology for each

    Each = stratified by <10%, 10-50%, 50-90%, 90+%

    DEPRECATED: makes tables that are too dense.
    """
    CSA_pure = df.loc[df['BaseDx'] == "Pure CSA"]
    CSA_predom = df.loc[df['BaseDx'] == "Predominantly CSA"]
    OSA_predom = df.loc[df['BaseDx'] == "Combined OSA/CSA"]
    OSA_pure = df.loc[df['BaseDx'] == "Mainly OSA"]

    num_total = len(df.index)
    num_csa_pure = len(CSA_pure.index)
    num_csa_predom = len(CSA_predom.index)
    num_osa_predom = len(OSA_predom.index)
    num_osa_pure = len(OSA_pure.index)

    column_labels = ['All, N=%s' % num_total,
                     'Pure CSA (90+%% CSA), N=%s' % num_csa_pure,
                     'Predominantly CSA (50-90%% CSA), N=%s' % num_csa_predom,
                     'Combined OSA/CSA (10-49.9%% CSA), N=%s' % num_osa_predom,
                     'Mainly OSA (<10%% CSA), N=%s' % num_osa_pure]

    workbook = Workbook()

    # Table 1: Demographics - each list is row

    demo_row_labels = ['Age',
                       'Sex',
                       'Ethnicity',
                       'Smoking Status',
                       'BMI',
                       'Comorbidity Counts',
                       'AHI']

    demographics = [(std_string(df['Age'].describe()), std_string(CSA_pure['Age'].describe()),
                     std_string(CSA_predom['Age'].describe()), std_string(OSA_predom['Age'].describe()),
                     std_string(OSA_pure['Age'].describe())),
                    (count_string(df['Sex'].value_counts(), num_total),
                     count_string(CSA_pure['Sex'].value_counts(), num_csa_pure),
                     count_string(CSA_predom['Sex'].value_counts(), num_csa_predom),
                     count_string(OSA_predom['Sex'].value_counts(), num_osa_predom),
                     count_string(OSA_pure['Sex'].value_counts(), num_osa_pure)),
                    (count_string(df['Race'].value_counts(), num_total),
                     count_string(CSA_pure['Race'].value_counts(), num_csa_pure),
                     count_string(CSA_predom['Race'].value_counts(), num_csa_predom),
                     count_string(OSA_predom['Race'].value_counts(), num_osa_predom),
                     count_string(OSA_pure['Race'].value_counts(), num_osa_pure)),
                    (count_string(df['Smoking'].value_counts(), num_total),
                     count_string(CSA_pure['Smoking'].value_counts(), num_csa_pure),
                     count_string(CSA_predom['Smoking'].value_counts(), num_csa_predom),
                     count_string(OSA_predom['Smoking'].value_counts(), num_osa_predom),
                     count_string(OSA_pure['Smoking'].value_counts(), num_osa_pure)),
                    (iqr_string(df['BMI'].describe()), iqr_string(CSA_pure['BMI'].describe()),
                     iqr_string(CSA_predom['BMI'].describe()), iqr_string(OSA_predom['BMI'].describe()),
                     iqr_string(OSA_pure['BMI'].describe())),
                    (count_string(histo_comorbs_includes(df), num_total),
                     count_string(histo_comorbs_includes(CSA_pure), num_csa_pure),
                     count_string(histo_comorbs_includes(CSA_predom), num_csa_predom),
                     count_string(histo_comorbs_includes(OSA_predom), num_osa_predom),
                     count_string(histo_comorbs_includes(OSA_pure), num_osa_pure)),
                    (iqr_string(df['AHI'].describe()), iqr_string(CSA_pure['AHI'].describe()),
                     iqr_string(CSA_predom['AHI'].describe()), iqr_string(OSA_predom['AHI'].describe()),
                     iqr_string(OSA_pure['AHI'].describe()))]

    demographics_df = pd.DataFrame(demographics, columns=column_labels, index=demo_row_labels)

    demographic_worksheet = workbook.worksheets[0]
    demographic_worksheet.title = "Demographics"

    for r in dataframe_to_rows(demographics_df, index=True, header=True):
        demographic_worksheet.append(r)

    for cell in demographic_worksheet['A'] + demographic_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')

    # Table 2 - Etiology
    etio_row_labels = ['Cause of Central Sleep Apnea', 'Heart Comorbidity Counts', 'CNS Comorbidity Counts']

    etiology = [
        (count_string(histo_dx_includes(df), num_total), count_string(histo_dx_includes(CSA_pure), num_csa_pure),
         count_string(histo_dx_includes(CSA_predom), num_csa_predom),
         count_string(histo_dx_includes(OSA_predom), num_osa_predom),
         count_string(histo_dx_includes(OSA_pure), num_osa_pure)),
        (count_string(histo_heart_includes(df), num_total), count_string(histo_heart_includes(CSA_pure), num_csa_pure),
         count_string(histo_heart_includes(CSA_predom), num_csa_predom),
         count_string(histo_heart_includes(OSA_predom), num_osa_predom),
         count_string(histo_heart_includes(OSA_pure), num_osa_pure)),
        (count_string(histo_cns_includes(df), num_total), count_string(histo_cns_includes(CSA_pure), num_csa_pure),
         count_string(histo_cns_includes(CSA_predom), num_csa_predom),
         count_string(histo_cns_includes(OSA_predom), num_osa_predom),
         count_string(histo_cns_includes(OSA_pure), num_osa_pure))]

    etiology_df = pd.DataFrame(etiology, columns=column_labels, index=etio_row_labels)

    etiology_worksheet = workbook.create_sheet(title="Etiology", index=1)

    for r in dataframe_to_rows(etiology_df, index=True, header=True):
        etiology_worksheet.append(r)

    for cell in etiology_worksheet['A'] + etiology_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center')

    #  do we want to include initial treatment? in the table 3? we're doing it

    # Table 3 - Outcome
    outcome_row_labels = ['Initial Treatment','Final Treatment', 'Outcome']

    outcome = [(count_string(df['InitTx'].value_counts(), num_total),
                count_string(CSA_pure['InitTx'].value_counts(), num_csa_pure),
                count_string(CSA_predom['InitTx'].value_counts(), num_csa_predom),
                count_string(OSA_predom['InitTx'].value_counts(), num_osa_predom),
                count_string(OSA_pure['InitTx'].value_counts(), num_osa_pure)),
               (count_string(df['FinalTx'].value_counts(), num_total),
                count_string(CSA_pure['FinalTx'].value_counts(), num_csa_pure),
                count_string(CSA_predom['FinalTx'].value_counts(), num_csa_predom),
                count_string(OSA_predom['FinalTx'].value_counts(), num_osa_predom),
                count_string(OSA_pure['FinalTx'].value_counts(), num_osa_pure)),
               (count_string(df['Outcome'].value_counts(), num_total),
                count_string(CSA_pure['Outcome'].value_counts(), num_csa_pure),
                count_string(CSA_predom['Outcome'].value_counts(), num_csa_predom),
                count_string(OSA_predom['Outcome'].value_counts(), num_osa_predom),
                count_string(OSA_pure['Outcome'].value_counts(), num_osa_pure))]

    outcome_df = pd.DataFrame(outcome, columns=column_labels, index=outcome_row_labels)

    outcome_worksheet = workbook.create_sheet(title="Outcome", index=2)

    for r in dataframe_to_rows(outcome_df, index=True, header=True):
        outcome_worksheet.append(r)

    for cell in outcome_worksheet['A'] + outcome_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center')

    # Table 4 - Outcome by etiology

    neurologic_df = df.loc[df['PostDx'].str.contains("Neurologic")].sort_values(by='Outcome')
    cardiac_df = df.loc[df['PostDx'].str.contains("Cardiac")].sort_values(by='Outcome')
    medication_df = df.loc[df['PostDx'].str.contains("Medication")].sort_values(by='Outcome')
    tecsa_df = df.loc[df['PostDx'].str.contains("TECSA")].sort_values(by='Outcome')
    # osacsa_df = df.loc[df['PostDx'].str.contains("OSA-CSA")].sort_values(by='Outcome')
    primary_df = df.loc[df['PostDx'].str.contains("Primary")].sort_values(by='Outcome')

    num_neurologic = len(neurologic_df.index)
    num_cardiac = len(cardiac_df.index)
    num_medication = len(medication_df.index)
    num_tecsa = len(tecsa_df.index)
    # num_osacsa = len(osacsa_df.index)
    num_primary = len(primary_df.index)

    column_etio_labels = ['All, n=%s' % num_total,
                     'Neurologic Contributor, n=%s' % num_neurologic,
                     'Cardiac Contributor, n=%s' % num_cardiac,
                     'Medication Contributor, n=%s' % num_medication,
                     'Treatment Emergent, n=%s' % num_tecsa,
                     # 'OSA-associated Centrals, n=%s' % num_osacsa,
                     'Primary CSA, n=%s' % num_primary]

    outcome_etio_row_labels = ['Initial Treatment', 'Final Treatment', 'Outcome']

    outcome_etio = [(count_string(df['InitTx'].value_counts(), num_total),
                    count_string(neurologic_df['InitTx'].value_counts(), num_neurologic),
                    count_string(cardiac_df['InitTx'].value_counts(), num_cardiac),
                    count_string(medication_df['InitTx'].value_counts(), num_medication),
                    count_string(tecsa_df['InitTx'].value_counts(), num_tecsa),
                    # count_string(osacsa_df['InitTx'].value_counts(), num_osacsa),
                    count_string(primary_df['InitTx'].value_counts(), num_primary)),
                    (count_string(df['FinalTx'].value_counts(), num_total),
                    count_string(neurologic_df['FinalTx'].value_counts(), num_neurologic),
                    count_string(cardiac_df['FinalTx'].value_counts(), num_cardiac),
                    count_string(medication_df['FinalTx'].value_counts(), num_medication),
                    count_string(tecsa_df['FinalTx'].value_counts(), num_tecsa),
                    # count_string(osacsa_df['FinalTx'].value_counts(), num_osacsa),
                    count_string(primary_df['FinalTx'].value_counts(), num_primary)),
                    (count_string(df['Outcome'].value_counts(), num_total),
                    count_string(neurologic_df['Outcome'].value_counts(), num_neurologic),
                    count_string(cardiac_df['Outcome'].value_counts(), num_cardiac),
                    count_string(medication_df['Outcome'].value_counts(), num_medication),
                    count_string(tecsa_df['Outcome'].value_counts(), num_tecsa),
                    # count_string(osacsa_df['Outcome'].value_counts(), num_osacsa),
                    count_string(primary_df['Outcome'].value_counts(), num_primary))]

    outcome_etio_df = pd.DataFrame(outcome_etio, columns=column_etio_labels, index=outcome_etio_row_labels)

    outcome_etio_worksheet = workbook.create_sheet(title="Outcome by Etiology", index=3)

    for r in dataframe_to_rows(outcome_etio_df, index=True, header=True):
        outcome_etio_worksheet.append(r)

    for cell in outcome_etio_worksheet['A'] + outcome_etio_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center')

    workbook.save("tables.xlsx")
    return


def iqr_string(summary):
    """returns string of 'mean [IQR 25,75], n=_' when given a dataframe.describe() result
    for non-normal dist data"""
    output = "".join(['%.1f' % summary['mean'], " [IQR ", '%.1f, ' % summary['25%'], '%.1f]' % summary['75%']])
    return output


def std_string(summary):
    """returns string of 'mean +/- std, n=_' when given a dataframe.describe() result
    for normal dist data"""
    output = "".join(['%.1f' % summary['mean'], " (+/- ", '%.1f)' % summary['std']])
    return output


def count_string(counts_series, num_patients):
    """returns string of the counts of each from a dataframe.value_counts() result and num_patients, which is the total
    number of patients (not observations, e.g. in the case of patients with multiple comorbidities, so that percentages
    of the patients can be calculated"""
    output = ""
    for label in counts_series.keys():
        output += label + " = %.0f" % counts_series[label]
        percent = (counts_series[label] / num_patients) * 100
        output += ' (%.1f%%)\n' % percent
    return output[:-1]  # take off the final \n


def count_string_indiv(num, num_patients):
    """returns an string with the number and percentage of an individuals value"""
    output = "%.0f/" % num
    output += str(num_patients)
    if num_patients is not 0:
        percentage = (num / num_patients) * 100
    else:
        percentage = 0.0
    output += ' (%.1f%%)' % percentage
    return output


def pieChartBaseDx(df):
    plt.style.use('seaborn-whitegrid')
    fig = plt.figure()  # container object
    ax = plt.axes()  # the box we'll draw in

    dx_counts = df['BaseDx'].value_counts().sort_index()
    colors = ["#d6cb9c", "#9cc1ec", "#8fd9c8", "#e7aeca"]  # IWantHue fancy, light
    ax.pie(dx_counts, labels=dx_counts.index, autopct="%1.1f%%", startangle=0,
           colors=colors, wedgeprops={'edgecolor': 'black'})
    ax.axis('equal')
    ax.set_title("Patients Categorized by Percentage of Apneas of Central Origin")
    ax.legend(["<10% Central", "10-50% Central", "50-90% Central", ">90% Central"],
              loc='lower left', frameon=True)
    plt.tight_layout()

    fig.savefig('Pie Chart Base Diagnosis.png', dpi=100)
    # plt.show()


def sankeyTypeFinalTx(df):
    fig = plt.figure()  # container object
    ax = plt.axes()  # the box we'll draw in

    dx_counts = df['BaseDx'].value_counts().sort_index()
    outcome_counts = df['FinalTx'].value_counts() * -1.0
    x = dx_counts.get_values().sum()

    flow = dx_counts.get_values().tolist() + outcome_counts.get_values().tolist()
    label = dx_counts.index.tolist() + outcome_counts.index.tolist()

    sk = sankey.Sankey(ax, head_angle=120, offset=0.4, scale=1 / float(x),
                       unit=" Pt", gap=1.0, margin=0.1,
                       flows=flow,
                       labels=label,
                       orientations=[1, 1, 0, -1, 1, 0, 1, -1, -1, -1, 1, -1, 1])

    # sk.add(flows=[0.05, 0.05, 0.9, -0.20, -0.15, -0.05, -0.50, -0.10],
    #    labels=['In1', 'In2', 'In3', 'First', 'Second', 'Third', 'Fourth', 'Fifth'],
    #    orientations=[-1, 1, 0, 1, 1, 1, 0, -1])

    sk.finish()
    # plt.tight_layout()
    ax.set_title("Percentage Central Apnea and FinalTx of Entire Dataset")
    ax.set_axis_off()

    fig.savefig('Sankey Type Treatment.png', dpi=100)
    # plt.show()


def sankeyTypeOutcome(df):
    fig = plt.figure()  # container object
    ax = plt.axes()  # the box we'll draw in

    dx_counts = df['BaseDx'].value_counts().sort_index()
    outcome_counts = df['Outcome'].value_counts() * -1.0
    x = dx_counts.get_values().sum()

    flow = dx_counts.get_values().tolist() + outcome_counts.get_values().tolist()
    label = dx_counts.index.tolist() + outcome_counts.index.tolist()

    sk = sankey.Sankey(ax, head_angle=120, offset=0.25, scale=1 / float(x),
                       unit=" patients", gap=0.9, margin=0.05,
                       flows=flow,
                       labels=label,
                       orientations=[1, 0, -1, -1, 1, -1, 0, -1, 1, -1])

    sk.finish()
    # plt.tight_layout()
    ax.set_title("Percentage Central Apnea and Outcome of Entire Dataset")
    ax.set_axis_off()

    fig.savefig('Sankey Type Outcome.png', dpi=100)
    # plt.show()


def vis_hist_etio(df):
    post_dx_histo = histo_dx_includes(df)
    hist_df = pd.DataFrame({"Dx": post_dx_histo.index, "Count": post_dx_histo.values})
    # print(hist_df)
    # hist_df = hist_df.drop(1)
    sns.set()
    sns.set_palette("husl", 3)
    ax = sns.catplot(x="Dx", y="Count", data=hist_df, kind='bar')
    plt.title("Etiology of CSA; Multiple Contributors Allowed")
    ax.set_axis_labels("Etiologies of Central Sleep Apnea", "Number of Patients")
    plt.legend()
    ax.fig.savefig('Histo Etiology.png', dpi=100)
    # plt.show()


def sankeyEtioTx(df):
    SMALL_SIZE = 8
    MEDIUM_SIZE = 10
    BIGGER_SIZE = 14

    plt.rc('font', size=SMALL_SIZE)  # controls default text sizes
    plt.rc('axes', titlesize=SMALL_SIZE)  # fontsize of the axes title
    plt.rc('axes', labelsize=MEDIUM_SIZE)  # fontsize of the x and y labels
    plt.rc('xtick', labelsize=SMALL_SIZE)  # fontsize of the tick labels
    plt.rc('ytick', labelsize=SMALL_SIZE)  # fontsize of the tick labels
    plt.rc('legend', fontsize=SMALL_SIZE)  # legend fontsize
    plt.rc('figure', titlesize=BIGGER_SIZE)  # fontsize of the figure title

    # fig, axs = plt.subplots(4,2)
    fig = plt.figure()
    spec = gridspec.GridSpec(ncols=2, nrows=5, figure=fig)

    f_ax1 = fig.add_subplot(spec[:-3, :])
    # f_ax2 = fig.add_subplot(spec[0, 1])
    f_ax3 = fig.add_subplot(spec[2, 0])
    f_ax4 = fig.add_subplot(spec[2, 1])
    f_ax5 = fig.add_subplot(spec[3, 0])
    f_ax6 = fig.add_subplot(spec[3, 1])
    f_ax7 = fig.add_subplot(spec[4, 0])
    f_ax8 = fig.add_subplot(spec[4, 1])

    fig.set_size_inches(18, 11)
    fig.suptitle(
        "Flow (Sankey) Diagram of Etiology of Central Apneas and Final Treatment, Separated by %CSA")
    sankeySubPlot(f_ax1, df, "All Patients Diagnosed with CSA")

    # f_ax3.set_title("Mainly OSA", fontsize=10)
    sankeySubPlot(f_ax3, df.loc[df['BaseDx'] == "Mainly OSA"],
                  "<10% CSAs")
    # f_ax5.set_title("Combined OSA/CSA", fontsize=10)
    sankeySubPlot(f_ax5, df.loc[df['BaseDx'] == "Combined OSA/CSA"],
                  "10-49.9% CSAs")
    # f_ax4.set_title("Predominantly CSA", fontsize=10)
    sankeySubPlot(f_ax4, df.loc[df['BaseDx'] == "Predominantly CSA"],
                  "50-90% CSAs")
    # f_ax6.set_title("Pure CSA", fontsize=10)
    sankeySubPlot(f_ax6, df.loc[df['BaseDx'] == "Pure CSA"],
                  ">90% CSAs")

    # Make Legend + Abbreviations
    sankeyLegendPlot(f_ax7)

    f_ax8.set_title("Abbreviations", fontsize=10)
    style = dict(size=7, color='black')
    f_ax8.text(0.1, -0.1,
               "Etiologies\n\
     Prim = Primary CSA\n\
     OSA = Central events associated only with Obstructive events\n\
     TE = Treatment Emergent CSA\n\
     CV = CSA associated with HFrEF, HFpEF, or AFib\n\
     CNS = CSA associated with CVA, TBI, Mass Lesion, Dementia, or Neurodegenerative disease\n\
     Med = CSA associated with high-dose opiate use\n\
     \n\
Treatments\n\
     Final tx ASV = patients eventually treated using Resmed/Respironics ASV, broken into flows by prior traetment tried\n\
     O2 = treatment with supplemental oxygen only (bleed-in not included)\n",
               ha="left", **style)
    f_ax8.set_axis_off()
    fig.tight_layout(rect=[0, 0.03, 1, 0.95])
    fig.savefig('Sankey Etio Treatment.png', dpi=100)
    # plt.show()


def sankeyLegendPlot(ax):
    ax.set_title("Legend", fontsize=10)
    sk = sankey.Sankey(ax=ax, scale=0.01, offset=0.3, head_angle=100,
                       unit=' Patients')
    sk.add(flows=[25, 5, 60, -35, -25, -40],
           labels=['Etiology 1', 'Etiology 2', 'Etiology 3 (Most Common)', 'Treatment 1', 'Treatment 2',
                   'Treatment 3 (most common)'],
           orientations=[-1, 1, 0, 1, -1, 0],
           pathlengths=[0.25, 0.25, 0.25, 0.25, 0.25, 0.25],
           patchlabel="All Patients\nn = 100\nRelative arrow size corresponds \nto proportion of patients",
           trunklength=5.0,
           facecolor='lightblue',
           alpha=0.5)  # Arguments to matplotlib.patches.PathPatch()
    sk.finish()
    ax.set_axis_off()

    return ax


def sankeySubPlot(ax, df, title):
    dx_counts = df['PostDx'].value_counts()
    dx_counts = dx_counts[dx_counts > 0]  # Drop labels with 0
    outcome_counts = df['FinalTx'].value_counts() * -1.0
    outcome_counts = outcome_counts[outcome_counts < 0]  # drop labels with 0
    x = dx_counts.get_values().sum()
    # print(dx_counts)
    # print(outcome_counts)

    flow = dx_counts.get_values().tolist() + outcome_counts.get_values().tolist()
    label = dx_counts.index.tolist() + outcome_counts.index.tolist()
    orientation = [0]  # Generate alternating 1, -1 for inputs / outputs
    pathlength = [0.25]
    for i in range(len(flow) - 1):
        if i == len(dx_counts) - 1:  # 1st outcome
            orientation.append(0)
            pathlength.append(0.25)
        else:
            if i % 2 == 1:
                orientation.append(1)
                if (i // 2) % 2 == 0:  # alternate (per side) pathlengths
                    pathlength.append(0.85)
                else:
                    pathlength.append(0.15)
            else:
                orientation.append(-1)
                if (i // 2) % 2 == 0:
                    pathlength.append(0.85)
                else:
                    pathlength.append(0.15)
    asvIndex = label.index("asv")
    label[asvIndex] = ""  # because this will be connected to sub-sankey

    sk = sankey.Sankey(ax, head_angle=120, offset=0.3, scale=1 / float(x),
                       unit="", gap=1.6, margin=0.05)
    sk.add(flows=flow,
           labels=label,
           orientations=orientation,
           pathlengths=pathlength,
           patchlabel=title + "\nn = " + str(x) + " patients",
           trunklength=6.0,
           facecolor='lightblue',
           alpha=0.75)

    # Create ASV subdigram
    ASV_df = df.loc[df['FinalTx'] == "asv"]
    ASV_dx_count = abs(flow[asvIndex])
    ASV_path_counts = ASV_df['ProcToASV'].value_counts() * -1.0
    ASV_flow = [ASV_dx_count]
    ASV_flow = ASV_flow + ASV_path_counts.tolist()
    ASV_label = [""]
    ASV_label = ASV_label + ASV_path_counts.index.tolist()
    ASV_orientation = [orientation[asvIndex] * -1, 0]  # start w/ ASV, opp
    ASV_pathlength = [1.0, 0.25]  # largest outcome straight

    for i in range(len(ASV_flow) - 2):
        if i % 2 == 1:
            ASV_orientation.append(1)
            if (i // 2) % 2 == 1:  # alternate (per side) pathlengths
                ASV_pathlength.append(0.5)
            else:
                ASV_pathlength.append(0.15)
        else:
            ASV_orientation.append(-1)
            if (i // 2) % 2 == 1:
                ASV_pathlength.append(0.5)
            else:
                ASV_pathlength.append(0.15)

    sk.add(flows=ASV_flow,
           labels=ASV_label,
           orientations=ASV_orientation,
           pathlengths=ASV_pathlength,
           prior=0,
           connect=(asvIndex, 0),
           patchlabel=" Final tx ASV",
           trunklength=4.5,
           facecolor='lavender',
           alpha=.75)

    sk.finish()
    ax.set_axis_off()
    return ax


def printSumByBaseDx(df):
    print("\n\n----AMONG THE ENTIRE DATASET----\n")
    summary_stats(df)

    print("\n\n----AMONG PATIENTS WITH PURE CSA (MORE THAN 90%)----\n")
    CSA_pure = df.loc[df['BaseDx'] == "Pure CSA"]
    summary_stats(CSA_pure)

    print("\n\n----AMONG PATIENTS WITH PREDOMINANTLY CSA (50 - 90%)----\n")
    CSA_predom = df.loc[df['BaseDx'] == "Predominantly CSA"]
    summary_stats(CSA_predom)

    print("\n\n----AMONG PATIENTS WITH Combined OSA/CSA (10 - 50% CSA)-----")
    OSA_predom = df.loc[df['BaseDx'] == "Combined OSA/CSA"]
    summary_stats(OSA_predom)

    print("\n\n----AMONG PATIENTS WITH PURE OSA (< 10% CSA)-----")
    OSA_pure = df.loc[df['BaseDx'] == "Mainly OSA"]
    summary_stats(OSA_pure)

    # print("\n\n----AMONG PATIENTS WITH PREDOMINANTLY CSA (MORE THAN 50%)----\n")
    # summary_stats(pd.merge(CSA_predom, CSA_pure, how='outer'))
    # print("\n\n----AMONG PATIENTS WITH < 50% CSA-----")
    # summary_stats(pd.merge(OSA_predom, OSA_pure, how='outer'))


def outcome_by_csa_percent(df):
    sns.set(style="white", palette=sns.color_palette("cubehelix", 6))

    f, axes = plt.subplots(4, 1, figsize=(6, 9), sharex=True)
    sns.despine(top=True, bottom=True)
    f.suptitle("Initial Outcome, Grouped by %CSA")

    CSA_pure = df.loc[df['BaseDx'] == "Pure CSA"]
    CSA_predom = df.loc[df['BaseDx'] == "Predominantly CSA"]
    OSA_predom = df.loc[df['BaseDx'] == "Combined OSA/CSA"]
    OSA_pure = df.loc[df['BaseDx'] == "Mainly OSA"]

    sns.countplot(y='Outcome', data=OSA_pure, ax=axes[0])
    axes[0].set(xlabel="", ylabel="<10% CSA")

    sns.countplot(y='Outcome', data=OSA_predom, ax=axes[1])
    axes[1].set(xlabel="", ylabel="10-49.9% CSA")

    sns.countplot(y='Outcome', data=CSA_predom, ax=axes[2])
    axes[2].set(xlabel="", ylabel="50-90% CSA")

    sns.countplot(y='Outcome', data=CSA_pure, ax=axes[3])
    axes[3].set(xlabel="Patients with each outcome after initial treatment", ylabel=">90% CSA")

    f.tight_layout(rect=[0, 0, 1, 0.95])
    f.savefig('Outcome by CSA percent.png', dpi=100)
    # plt.show()
    return


def outcome_by_etio(df):
    sns.set(style="white", palette=sns.color_palette("cubehelix", 6)) #still 6 if osa csa
    f, axes = plt.subplots(5, 2, figsize=(9, 9)) # 6, 2 if OSA CSA
    sns.despine(top=True, bottom=True)
    f.suptitle("Outcome, Grouped by Contributing Etiology")

    # contains used instead of equal to include patients with multiple etio (e.g. cardiac+medication count to both)
    neurologic_df = df.loc[df['PostDx'].str.contains("Neurologic")].sort_values(by='Outcome')
    cardiac_df = df.loc[df['PostDx'].str.contains("Cardiac")].sort_values(by='Outcome')
    medication_df = df.loc[df['PostDx'].str.contains("Medication")].sort_values(by='Outcome')
    tecsa_df = df.loc[df['PostDx'].str.contains("TECSA")].sort_values(by='Outcome')
    # osacsa_df = df.loc[df['PostDx'].str.contains("OSA-CSA")].sort_values(by='Outcome')
    primary_df = df.loc[df['PostDx'].str.contains("Primary")].sort_values(by='Outcome')

    # Create count plot for each Etio on the left, then a Pie Chart with proportion on the right

    # Neurologic
    sns.countplot(y='Outcome', data=neurologic_df, ax=axes[0,0])
    axes[0,0].set(xlabel="", ylabel="Neurologic")
    neuro_counts = neurologic_df['Outcome'].value_counts().sort_index()
    neuro_wedges, _, _ = axes[0, 1].pie(neuro_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
                                           textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                           wedgeprops={'edgecolor': 'black'})
    axes[0, 1].legend(neuro_wedges, neuro_counts.index, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')

    # Cardiac
    sns.countplot(y='Outcome', data=cardiac_df, ax=axes[1,0])
    axes[1,0].set(xlabel="", ylabel="Cardiac")
    cardiac_counts = cardiac_df['Outcome'].value_counts().sort_index()
    cardiac_wedges, _, _ = axes[1, 1].pie(cardiac_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
                                           textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                           wedgeprops={'edgecolor': 'black'})
    axes[1, 1].legend(cardiac_wedges, cardiac_counts.index, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')

    # Medication
    sns.countplot(y='Outcome', data=medication_df, ax=axes[2,0])
    axes[2,0].set(xlabel="", ylabel="Medication")
    medication_counts = medication_df['Outcome'].value_counts().sort_index()
    medication_wedges, _, _ = axes[2, 1].pie(medication_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
                                           textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                           wedgeprops={'edgecolor': 'black'})
    axes[2, 1].legend(medication_wedges, medication_counts.index, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')

    # OSA-CSA
    # sns.countplot(y='Outcome', data=osacsa_df, ax=axes[3,0])
    # axes[3,0].set(xlabel="", ylabel="OSA-CSA")
    # osacsa_counts = osacsa_df['Outcome'].value_counts().sort_index()
    # osacsa_wedges, _, _ = axes[3, 1].pie(osacsa_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
    #                                        textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
    #                                        wedgeprops={'edgecolor': 'black'})
    # axes[3, 1].legend(osacsa_wedges, osacsa_counts.index, loc="center left",
    #                   bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')

    # If adding OSA-CSA back, would need to increase by 1 all of the axes indices

    # TE-CSA
    sns.countplot(y='Outcome', data=tecsa_df, ax=axes[3,0])
    axes[3,0].set(xlabel="", ylabel="TE-CSA")
    tecsa_counts = tecsa_df['Outcome'].value_counts().sort_index()
    tecsa_wedges, _, _ = axes[3, 1].pie(tecsa_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
                                           textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                           wedgeprops={'edgecolor': 'black'})
    axes[3, 1].legend(tecsa_wedges, tecsa_counts.index, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')

    #Primary
    sns.countplot(y='Outcome', data=primary_df, ax=axes[4,0])
    axes[4,0].set(xlabel="Outcome of initial treatment by etiology", ylabel="Primary CSA")
    primary_counts = primary_df['Outcome'].value_counts().sort_index()
    primary_wedges, _, _ = axes[4, 1].pie(primary_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
                                           textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                           wedgeprops={'edgecolor': 'black'})
    axes[4, 1].legend(primary_wedges, primary_counts.index, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')
    axes[4, 1].set(xlabel="\nProportion with each outcome\nby etiology")

    # Combined X axis for L side
    axes[4, 0].get_shared_x_axes().join(axes[4, 0], axes[3, 0], axes[2, 0], axes[1, 0], axes[0, 0]) # axes[5, 0] would need to be added back
    axes[0, 0].set_xticklabels("")
    axes[1, 0].set_xticklabels("")
    axes[2, 0].set_xticklabels("")
    axes[3, 0].set_xticklabels("")
    # axes[4, 0].set_xticklabels("")
    # Leave bottom aka [5,0] labels in

    # Resize all
    axes[0, 0].autoscale()
    axes[1, 0].autoscale()
    axes[2, 0].autoscale()
    axes[3, 0].autoscale()
    axes[4, 0].autoscale()
    # axes[5, 0].autoscale()

    f.tight_layout(rect=[0, 0, 1, 0.95])
    f.savefig('Outcome by Etio.png', dpi=100)
    # plt.show()


def outcome_by_etio_no_pie(df):
    """ version of the outcome graph with just bars, no pie
    """

    # TODO: possibly combine with etio by percentage

    sns.set(style="white", palette=sns.color_palette("cubehelix", 6)) #still 6 if osa csa
    f, axes = plt.subplots(5, 1, figsize=(6, 9)) # 6, 2 if OSA CSA
    sns.despine(top=True, bottom=True)
    # f.suptitle("Outcome, Grouped by Contributing Etiology")

    # contains used instead of equal to include patients with multiple etio (e.g. cardiac+medication count to both)
    neurologic_df = df.loc[df['PostDx'].str.contains("Neurologic")].sort_values(by='Outcome')
    cardiac_df = df.loc[df['PostDx'].str.contains("Cardiac")].sort_values(by='Outcome')
    medication_df = df.loc[df['PostDx'].str.contains("Medication")].sort_values(by='Outcome')
    tecsa_df = df.loc[df['PostDx'].str.contains("TECSA")].sort_values(by='Outcome')
    # osacsa_df = df.loc[df['PostDx'].str.contains("OSA-CSA")].sort_values(by='Outcome')
    primary_df = df.loc[df['PostDx'].str.contains("Primary")].sort_values(by='Outcome')

    # collapse possible outcomes
    neurologic_df['col_outcome'] = neurologic_df.apply(collapse_initial_outcome, axis=1)
    cardiac_df['col_outcome'] = cardiac_df.apply(collapse_initial_outcome, axis=1)
    medication_df['col_outcome'] = medication_df.apply(collapse_initial_outcome, axis=1)
    tecsa_df['col_outcome'] = tecsa_df.apply(collapse_initial_outcome, axis=1)
    # osacsa_df['col_outcome'] = osacsa_df.apply(collapse_initial_outcome, axis=1)
    primary_df['col_outcome'] = primary_df.apply(collapse_initial_outcome, axis=1)

    # Create count plot for each Etio on the left, then a Pie Chart with proportion on the right

    hatches = ['///', '|||', 'xxx', '\\\\\\', '', '+++']
    face_color = ['dimgray', 'silver', 'whitesmoke', 'grey', 'gainsboro', 'darkgrey']

    # Neurologic
    bar = sns.countplot(y='col_outcome', data=neurologic_df, ax=axes[0])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[0].set(xlabel="", ylabel="Neurologic")


    # Cardiac
    bar = sns.countplot(y='col_outcome', data=cardiac_df, ax=axes[1])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[1].set(xlabel="", ylabel="Cardiac")

    # Medication
    bar = sns.countplot(y='col_outcome', data=medication_df, ax=axes[2])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[2].set(xlabel="", ylabel="Medication")

    # OSA-CSA
    # bar = sns.countplot(y='col_outcome', data=osacsa_df, ax=axes[3,0])
    # for i, this_bar in enumerate(bar.patches):
    #     # Set a different hatch for each bar
    #     this_bar.set_hatch(hatches[i])
    # axes[3].set(xlabel="", ylabel="OSA-CSA")
    # If adding OSA-CSA back, would need to increase by 1 all of the axes indices

    # TE-CSA
    bar = sns.countplot(y='col_outcome', data=tecsa_df, ax=axes[3])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[3].set(xlabel="", ylabel="TE-CSA")

    #Primary
    bar = sns.countplot(y='col_outcome', data=primary_df, ax=axes[4])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor('white')
        this_bar.set_facecolor(face_color[i])
    axes[4].set(xlabel="Outcome of initial treatment by etiology", ylabel="Primary CSA")

    # Combined X axis for L side
    axes[4].get_shared_x_axes().join(axes[4], axes[3], axes[2], axes[1], axes[0]) # axes[5] would need to be added back
    axes[0].set_xticklabels("")
    axes[1].set_xticklabels("")
    axes[2].set_xticklabels("")
    axes[3].set_xticklabels("")
    # axes[4].set_xticklabels("")
    # Leave bottom labels in

    # Resize all
    axes[0].autoscale()
    axes[1].autoscale()
    axes[2].autoscale()
    axes[3].autoscale()
    axes[4].autoscale()
    # axes[5].autoscale()

    f.tight_layout(rect=[0, 0, 1, 1])
    f.savefig('Outcome by Etio no pie.png', dpi=100)
    # plt.show()


def init_tx_by_etio(df):
    sns.set(style="white", palette=sns.color_palette("cubehelix", 6))

    f, axes = plt.subplots(5, 1, figsize=(6, 9), sharex=True) # 6 if osa csa
    sns.despine(top=True, bottom=True)
    f.suptitle("Initial Treatment Modality, Grouped by Contributing Etiology")

    # contains used instead of equal to include patients with multiple etio (e.g. cardiac+medication count to both)
    neurologic_df = df.loc[df['PostDx'].str.contains("Neurologic")]
    cardiac_df = df.loc[df['PostDx'].str.contains("Cardiac")]
    medication_df = df.loc[df['PostDx'].str.contains("Medication")]
    tecsa_df = df.loc[df['PostDx'].str.contains("TECSA")]
    # osacsa_df = df.loc[df['PostDx'].str.contains("OSA-CSA")]
    primary_df = df.loc[df['PostDx'].str.contains("Primary")]

    sns.countplot(y='InitTx', data=neurologic_df, ax=axes[0])
    axes[0].set(xlabel="", ylabel="Neurologic")

    sns.countplot(y='InitTx', data=cardiac_df, ax=axes[1])
    axes[1].set(xlabel="", ylabel="Cardiac")

    sns.countplot(y='InitTx', data=medication_df, ax=axes[2])
    axes[2].set(xlabel="", ylabel="Medication")

    # sns.countplot(y='InitTx', data=osacsa_df, ax=axes[3])
    # axes[3].set(xlabel="", ylabel="OSA-CSA")

    sns.countplot(y='InitTx', data=tecsa_df, ax=axes[3])
    axes[3].set(xlabel="", ylabel="TE-CSA")

    sns.countplot(y='InitTx', data=primary_df, ax=axes[4])
    axes[4].set(xlabel="Patients who initially received each treatment modality", ylabel="Primary CSA")

    f.tight_layout(rect=[0, 0, 1, 0.95])
    f.savefig('Init Treatment by Etiology.png', dpi=100)
    # plt.show()


def test_by_etio(df):
    sns.set(style="white", palette=sns.color_palette("cubehelix", 2))

    f, axes = plt.subplots(5, 1, figsize=(4, 9), sharex=True) # 6 if osa csa
    sns.despine(top=True, bottom=True)
    f.suptitle("Diagnostic Test\nGrouped by Contributing Etiology")

    # contains used instead of equal to include patients with multiple etio (e.g. cardiac+medication count to both)
    neurologic_df = df.loc[df['PostDx'].str.contains("Neurologic")].sort_values(by='StudyType')
    cardiac_df = df.loc[df['PostDx'].str.contains("Cardiac")].sort_values(by='StudyType')
    medication_df = df.loc[df['PostDx'].str.contains("Medication")].sort_values(by='StudyType')
    tecsa_df = df.loc[df['PostDx'].str.contains("TECSA")].sort_values(by='StudyType')
    # osacsa_df = df.loc[df['PostDx'].str.contains("OSA-CSA")].sort_values(by='Outcome')
    primary_df = df.loc[df['PostDx'].str.contains("Primary")].sort_values(by='StudyType')

    # Create count plot for each Etio on the left, then a Pie Chart with proportion on the right

    # Neurologic
    axes[0].set(xlabel="", ylabel="Neurologic")
    neuro_counts = neurologic_df['StudyType'].value_counts().sort_index()
    neuro_wedges, _, _ = axes[0].pie(neuro_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
                                        textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                        wedgeprops={'edgecolor': 'black'})
    axes[0].legend(neuro_wedges, neuro_counts.index, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')

    # Cardiac
    axes[1].set(xlabel="", ylabel="Cardiac")
    cardiac_counts = cardiac_df['StudyType'].value_counts().sort_index()
    cardiac_wedges, _, _ = axes[1].pie(cardiac_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
                                          textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                          wedgeprops={'edgecolor': 'black'})
    axes[1].legend(cardiac_wedges, cardiac_counts.index, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')

    # Medication
    axes[2].set(xlabel="", ylabel="Medication")
    medication_counts = medication_df['StudyType'].value_counts().sort_index()
    medication_wedges, _, _ = axes[2].pie(medication_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
                                             textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                             wedgeprops={'edgecolor': 'black'})
    axes[2].legend(medication_wedges, medication_counts.index, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')

    # OSA-CSA
    # axes[3].set(xlabel="", ylabel="OSA-CSA")
    # osacsa_counts = osacsa_df['Outcome'].value_counts().sort_index()
    # osacsa_wedges, _, _ = axes[3].pie(osacsa_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
    #                                        textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
    #                                        wedgeprops={'edgecolor': 'black'})
    # axes[3].legend(osacsa_wedges, osacsa_counts.index, loc="center left",
    #                   bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')

    # If adding OSA-CSA back, would need to increase by 1 all of the axes indices

    # TE-CSA
    axes[3].set(xlabel="", ylabel="TE-CSA")
    tecsa_counts = tecsa_df['StudyType'].value_counts().sort_index()
    tecsa_wedges, _, _ = axes[3].pie(tecsa_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
                                        textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                        wedgeprops={'edgecolor': 'black'})
    axes[3].legend(tecsa_wedges, tecsa_counts.index, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')

    # Primary
    axes[4].set(xlabel="Outcome of initial treatment by etiology", ylabel="Primary CSA")
    primary_counts = primary_df['StudyType'].value_counts().sort_index()
    primary_wedges, _, _ = axes[4].pie(primary_counts, autopct="%1.1f%%", startangle=0, pctdistance=1.35,
                                          textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                          wedgeprops={'edgecolor': 'black'})
    axes[4].legend(primary_wedges, primary_counts.index, loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1), fontsize='x-small')
    axes[4].set(xlabel="\nProportion using each diagnostic \ntest, grouped by etiology")

    f.tight_layout(rect=[0, 0, 1, 0.95])
    f.savefig('Diag Test by Etiology.png', dpi=100)
    # plt.show()


def init_tx_by_csa(df):
    sns.set(style="white", palette=sns.color_palette("cubehelix", 6))

    f, axes = plt.subplots(4, 1, figsize=(6, 9), sharex=True)
    sns.despine(top=True, bottom=True)
    f.suptitle("Initial Treatment Modality, Grouped by %CSA")

    CSA_pure = df.loc[df['BaseDx'] == "Pure CSA"]
    CSA_predom = df.loc[df['BaseDx'] == "Predominantly CSA"]
    OSA_predom = df.loc[df['BaseDx'] == "Combined OSA/CSA"]
    OSA_pure = df.loc[df['BaseDx'] == "Mainly OSA"]

    sns.countplot(y='InitTx', data=OSA_pure, ax=axes[0])
    axes[0].set(xlabel="", ylabel="<10% CSA")

    sns.countplot(y='InitTx', data=OSA_predom, ax=axes[1])
    axes[1].set(xlabel="", ylabel="10-49.9% CSA")

    sns.countplot(y='InitTx', data=CSA_predom, ax=axes[2])
    axes[2].set(xlabel="", ylabel="50-90% CSA")

    sns.countplot(y='InitTx', data=CSA_pure, ax=axes[3])
    axes[3].set(xlabel="Patients who initially received each treatment modality", ylabel=">90% CSA")

    f.tight_layout(rect=[0, 0, 1, 0.95])
    f.savefig('Initial Treatment by Perc CSA.png', dpi=100)
    # plt.show()
    return


def final_tx_by_csa(df):
    sns.set(style="white", palette=sns.color_palette("cubehelix", 6))

    f, axes = plt.subplots(4, 1, figsize=(6, 9), sharex=True)
    sns.despine(top=True, bottom=True)
    f.suptitle("Final Treatment Modality, Grouped by %CSA")

    CSA_pure = df.loc[df['BaseDx'] == "Pure CSA"]
    CSA_predom = df.loc[df['BaseDx'] == "Predominantly CSA"]
    OSA_predom = df.loc[df['BaseDx'] == "Combined OSA/CSA"]
    OSA_pure = df.loc[df['BaseDx'] == "Mainly OSA"]

    sns.countplot(y='FinalTx', data=OSA_pure, ax=axes[0])
    axes[0].set(xlabel="", ylabel="<10% CSA")

    sns.countplot(y='FinalTx', data=OSA_predom, ax=axes[1])
    axes[1].set(xlabel="", ylabel="10-49.9% CSA")

    sns.countplot(y='FinalTx', data=CSA_predom, ax=axes[2])
    axes[2].set(xlabel="", ylabel="50-90% CSA")

    sns.countplot(y='FinalTx', data=CSA_pure, ax=axes[3])
    axes[3].set(xlabel="Patients who initially received each treatment modality", ylabel=">90% CSA")

    f.tight_layout(rect=[0, 0, 1, 0.95])
    f.savefig('Final Treatment by Perc CSA.png', dpi=100)
    # plt.show()
    return


def test_by_csa(df):
    """creates a horizontal count chart with counts of each etiology and a pie chart with the proportion,
     grouped by percentage of CSA"""
    sns.set(style="white", palette=sns.color_palette("cubehelix", 3))
    f, axes = plt.subplots(4, 1, figsize=(4, 9))#, sharex=True)
    sns.despine(top=True, bottom=True)
    f.suptitle("Diagnostic Test\nGrouped by %CSA")

    OSA_pure_df = df.loc[df['BaseDx'] == "Mainly OSA"]
    OSA_predom_df = df.loc[df['BaseDx'] == "Combined OSA/CSA"]
    CSA_predom_df = df.loc[df['BaseDx'] == "Predominantly CSA"]
    CSA_pure_df = df.loc[df['BaseDx'] == "Pure CSA"]

    OSA_pure_hist = OSA_pure_df['StudyType'].value_counts()
    OSA_predom_hist = OSA_predom_df['StudyType'].value_counts()
    CSA_predom_hist = CSA_predom_df['StudyType'].value_counts()
    CSA_pure_hist = CSA_pure_df['StudyType'].value_counts()

    # Pure OSA
    axes[0].set(xlabel="", ylabel="<10% CSA")
    osa_pure_wedges, _, _ = axes[0].pie(OSA_pure_hist, autopct="%1.1f%%", startangle=0, pctdistance=1.25,
                                             textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                             wedgeprops={'edgecolor': 'black'})
    axes[0].legend(osa_pure_wedges, OSA_pure_hist.keys(), loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

    # Predom OSA
    axes[1].set(xlabel="", ylabel="10-49.9% CSA")
    osa_predom_wedges, _, _ = axes[1].pie(OSA_predom_hist, autopct="%1.1f%%", startangle=0, pctdistance=1.25,
                                             textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                             wedgeprops={'edgecolor': 'black'})
    axes[1].legend(osa_predom_wedges, OSA_predom_hist.keys(), loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

    # Predom CSA
    axes[2].set(xlabel="", ylabel="50-90% CSA")

    csa_predom_wedges, _, _ = axes[2].pie(CSA_predom_hist, autopct="%1.1f%%", startangle=0, pctdistance=1.25,
                                             textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                             wedgeprops={'edgecolor': 'black'})
    axes[2].legend(csa_predom_wedges, CSA_predom_hist.keys(), loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

    # Pure CSA
    axes[3].set(xlabel="Patients With Each Etiology Contributing to CSA", ylabel=">90% CSA")

    csa_pure_wedges, _, _ = axes[3].pie(CSA_pure_hist, autopct="%1.1f%%", startangle=0, pctdistance=1.25,
                                             textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                             wedgeprops={'edgecolor': 'black'})
    axes[3].legend(csa_pure_wedges, CSA_pure_hist.keys(), loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    axes[3].set(xlabel="\nProportion using each type \nof diagnostic test")

    f.tight_layout(rect=[0, 0, 1, 0.95]) # .95 to leave space for title
    f.savefig('Diag Test by percentage CSA.png', dpi=100)
    # plt.show()


def etio_by_csa(df):
    """creates a horizontal count chart with counts of each etiology and a pie chart with the proportion,
     grouped by percentage of CSA"""
    sns.set(style="white", palette=sns.color_palette("cubehelix", 6))
    f, axes = plt.subplots(4, 2, figsize=(9, 9))#, sharex=True)
    sns.despine(top=True, bottom=True)
    f.suptitle("Etiology of Central Events, Grouped by %CSA")

    OSA_pure_hist = histo_dx_includes(df.loc[df['BaseDx'] == "Mainly OSA"], return_df=True).sort_values("Dx")
    OSA_predom_hist = histo_dx_includes(df.loc[df['BaseDx'] == "Combined OSA/CSA"], return_df=True).sort_values("Dx")
    CSA_predom_hist = histo_dx_includes(df.loc[df['BaseDx'] == "Predominantly CSA"], return_df=True).sort_values("Dx")
    CSA_pure_hist = histo_dx_includes(df.loc[df['BaseDx'] == "Pure CSA"], return_df=True).sort_values("Dx")

    # Create count plot for each #CSA on the left, then a Pie Chart with proportion on the right

    # Pure OSA
    sns.barplot(x="Count", y="Dx", data=OSA_pure_hist, ax=axes[0,0])
    axes[0, 0].set(xlabel="", ylabel="<10% CSA")
    osa_pure_wedges, _, _ = axes[0, 1].pie(OSA_pure_hist['Count'], autopct="%1.1f%%", startangle=0, pctdistance=1.25,
                                             textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                             wedgeprops={'edgecolor': 'black'})
    axes[0, 1].legend(osa_pure_wedges, OSA_pure_hist['Dx'], loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

    # Predom OSA
    sns.barplot(x="Count", y="Dx", data=OSA_predom_hist, ax=axes[1,0])
    axes[1, 0].set(xlabel="", ylabel="10-49.9% CSA")
    osa_predom_wedges, _, _ = axes[1, 1].pie(OSA_predom_hist['Count'], autopct="%1.1f%%", startangle=0, pctdistance=1.25,
                                             textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                             wedgeprops={'edgecolor': 'black'})
    axes[1, 1].legend(osa_predom_wedges, OSA_predom_hist['Dx'], loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

    # Predom CSA
    sns.barplot(x="Count", y="Dx", data=CSA_predom_hist, ax=axes[2, 0])
    axes[2, 0].set(xlabel="", ylabel="50-90% CSA")

    csa_predom_wedges, _, _ = axes[2, 1].pie(CSA_predom_hist['Count'], autopct="%1.1f%%", startangle=0, pctdistance=1.25,
                                             textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                             wedgeprops={'edgecolor': 'black'})
    axes[2, 1].legend(csa_predom_wedges, CSA_predom_hist['Dx'], loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

    # Pure CSA
    sns.barplot(x="Count", y="Dx", data=CSA_pure_hist, ax=axes[3,0])
    axes[3, 0].set(xlabel="Patients With Each Etiology Contributing to Central Events", ylabel=">90% CSA")

    csa_pure_wedges, _, _ = axes[3, 1].pie(CSA_pure_hist['Count'], autopct="%1.1f%%", startangle=0, pctdistance=1.25,
                                             textprops={'size': 'x-small'}, colors=sns.color_palette("cubehelix", 6),
                                             wedgeprops={'edgecolor': 'black'})
    axes[3, 1].legend(csa_pure_wedges, CSA_pure_hist['Dx'], loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    axes[3, 1].set(xlabel="\nProportion with each etiology\nContributing to Central Events")

    # Combined X axis for L side
    axes[3, 0].get_shared_x_axes().join(axes[3,0], axes[2,0], axes[1,0], axes[0,0])
    axes[0, 0].set_xticklabels("")
    axes[1, 0].set_xticklabels("")
    axes[2, 0].set_xticklabels("")
    # Leave bottom aka [3,0] labels in

    # Resize all
    axes[0, 0].autoscale()
    axes[1, 0].autoscale()
    axes[2, 0].autoscale()
    axes[3, 0].autoscale()

    f.tight_layout(rect=[0, 0, 1, 0.95]) # .95 to leave space for title
    f.savefig('Etio by percentage CSA.png', dpi=100)
    # plt.show()


def etio_by_csa_no_pie(df):
    """creates a horizontal count chart with counts of each etiology and a pie chart with the proportion,
     grouped by percentage of CSA"""

    # TODO: see if there's a way to combine this information with Outcome by Etio

    sns.set(style="white", palette=sns.color_palette("cubehelix", 6))
    f, axes = plt.subplots(4, 1, figsize=(6, 9))#, sharex=True)
    sns.despine(top=True, bottom=True)
    #f.suptitle("Etiology of Central Events, Grouped by %Central Events")

    OSA_pure_hist = histo_dx_includes(df.loc[df['BaseDx'] == "Mainly OSA"], return_df=True).sort_values("Dx")
    OSA_predom_hist = histo_dx_includes(df.loc[df['BaseDx'] == "Combined OSA/CSA"], return_df=True).sort_values("Dx")
    CSA_predom_hist = histo_dx_includes(df.loc[df['BaseDx'] == "Predominantly CSA"], return_df=True).sort_values("Dx")
    CSA_pure_hist = histo_dx_includes(df.loc[df['BaseDx'] == "Pure CSA"], return_df=True).sort_values("Dx")

    # Create count plot for each #CSA on the left, then a Pie Chart with proportion on the right

    hatches = ['///', '|||', 'xxx', '\\\\\\', '', '+++']
    face_color = [ 'dimgray', 'silver', 'whitesmoke', 'grey', 'gainsboro', 'darkgrey']

    # Pure OSA
    bar = sns.barplot(x="Count", y="Dx", data=OSA_pure_hist, ax=axes[3])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    #axes[3].set(xlabel="Patients With Each Etiology Contributing to Central Events", ylabel="<10% Central Events", yticklabels = [])
    axes[3].set(xlabel="Patients With Each Etiology Contributing to Central Events", ylabel="")

    # Predom OSA
    bar = sns.barplot(x="Count", y="Dx", data=OSA_predom_hist, ax=axes[2])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[2].set(xlabel="", ylabel="")
    # axes[2].set(xlabel="", ylabel="10-50% Central Events", yticklabels=[])

    # Predom CSA
    bar = sns.barplot(x="Count", y="Dx", data=CSA_predom_hist, ax=axes[1])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[1].set(xlabel="", ylabel="")
    # axes[1].set(xlabel="", ylabel="50-90% Central Events", yticklabels=[])

    # Pure CSA
    bar = sns.barplot(x="Count", y="Dx", data=CSA_pure_hist, ax=axes[0])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[0].set(xlabel="", ylabel="")
    # axes[0].set(xlabel="", ylabel=">90% Central Events", yticklabels=[])

    # Combined X axis for L side
    axes[3].get_shared_x_axes().join(axes[3], axes[2], axes[1], axes[0])
    axes[0].set_xticklabels("")
    axes[1].set_xticklabels("")
    axes[2].set_xticklabels("")
    # Leave bottom aka [3,0] labels in

    # Resize all
    axes[0].autoscale()
    axes[1].autoscale()
    axes[2].autoscale()
    axes[3].autoscale()

    f.tight_layout(rect=[0, 0, 1, 1]) # .95 to leave space for title
    f.savefig('Etio by percentage CSA no pie.png', dpi=100)
    # plt.show()


def etio_by_csa_dep(df):
    """creates a lollipop graph with counts of each etiology, grouped by percentage of CSA"""

    #TODO: need to scale subplots to same axes, spread subplots apart a little.
    #Note: plan is to replace this with some seaborn graphs that are easier.

    # fig, axs = plt.subplots(4,1)
    fig = plt.figure()
    spec = gridspec.GridSpec(ncols=1, nrows=4, figure=fig)

    f_ax1 = fig.add_subplot(spec[0, 0])
    f_ax2 = fig.add_subplot(spec[1, 0])
    f_ax3 = fig.add_subplot(spec[2, 0])
    f_ax4 = fig.add_subplot(spec[3, 0])

    OSA_pure = df.loc[df['BaseDx'] == "Mainly OSA"]
    OSA_predom = df.loc[df['BaseDx'] == "Combined OSA/CSA"]
    CSA_predom = df.loc[df['BaseDx'] == "Predominantly CSA"]
    CSA_pure = df.loc[df['BaseDx'] == "Pure CSA"]

    fig.suptitle("Contributing etiologies grouped by percentage central events")

    etio_subplot(OSA_pure, f_ax1, "Pure OSA", graph_color='forestgreen')
    etio_subplot(OSA_predom, f_ax2, "Predom OSA", graph_color='aquamarine')
    etio_subplot(CSA_predom, f_ax3, "Predom CSA", graph_color='lightskyblue')
    etio_subplot(CSA_pure, f_ax4, "Pure CSA", graph_color='slategray')

    plt.show()

def etio_subplot(df, ax, title, graph_color='skyblue'):
    """creates the lollipop plot on the given axes"""

    post_dx_histo = histo_dx_includes(df)
    hist_df = pd.DataFrame({"Dx": post_dx_histo.index, "Count": post_dx_histo.data})
    #hist_df = hist_df.drop(1)
    print(hist_df)

    graph_range = range(1,len(hist_df.index)+1)
    ax.hlines(y=graph_range, xmin=0, xmax=hist_df['Count'], color=graph_color)
    ax.plot(hist_df['Count'], graph_range, "D", color=graph_color)
    ax.set_yticks(range(1, len(hist_df['Dx'])+1))
    ax.set_yticklabels(hist_df['Dx'], fontsize='10')

    ax.set_title(title, fontsize='10')
    return ax

def visualizations(df):
    # For testing visualizations

    plt.style.use('seaborn-whitegrid')
    fig = plt.figure()  # container object
    ax = plt.axes()  # the box we'll draw in
    sns.set()
    sns.set_palette("husl",3)

    # Severity of OSA by percent CSAs and Sex
    # ax = sns.boxplot('BaseDx', 'AHI', hue='Sex', data=df)

    # Outcome by Base Dx as histograms
    # Todo: figure out how to normalize this
    ax = sns.catplot('BaseDx', hue='Outcome', kind='count', data=df)

    # Vis of distribution of AHIs
    #cleaned_df = df[df['AHI'].notnull()]
    #sns.distplot(cleaned_df['AHI'])

    plt.show()


def coded_output(database_df, output_loc='coded_output.xlsx'):
    """ takes the database dataframe as input
    makes an output excel with integer codes for value (as opposed to the categorical -
    e.g. male = 0, female = 1

    writes them to
    output_loc - for the encoded database
    keys_output_loc - for the keys

    Goal:
    ID no coding
    Age no coding
    Sex coded
    Race coded
    Smoking coded
    BMI no coding
    Comorbs broken into
    -   HTN T / F
    -   DM
    -   Psych
    -   Renal
    -   Heart?
    -   CNS
    -   Opiate
    AHI no coding
    HFrEF y/n
    HFpEF and AF  ([ ] is it possible? HFpEF vs HFrEF coded?)
    HFrEF and AF y/n
    Stroke y/n
    Dementia or neurodegen
    Dementia and stroke or neurodegen
    Final Treatment
    Category of OSA
    Dx Study

    """

    LE = LabelEncoder()
    output_df = pd.DataFrame()
    labels_ser = pd.Series()

    # ID no coding
    output_df["ID"] = database_df["ID"]
    labels_ser["ID"] = "no coding"

    # Age no coding
    output_df["Age"] = database_df['Age']
    labels_ser["Age"] = "no coding"

    # Sex coded
    LE.fit(database_df['Sex'])
    labels_ser['Sex'] = create_key_string(list(LE.classes_))
    output_df['Sex'] = LE.transform(database_df['Sex'])

    # Race coded
    LE.fit(database_df['Race'])
    labels_ser['Race'] = create_key_string(list(LE.classes_))
    output_df['Race'] = LE.transform(database_df['Race'])

    # Smoking coded
    LE.fit(database_df['Smoking'])
    labels_ser['Smoking'] = create_key_string(list(LE.classes_))
    output_df['Smoking'] = LE.transform(database_df['Smoking'])

    # BMI no coding
    output_df["BMI"] = database_df["BMI"]
    labels_ser["BMI"] = "no coding"

    # Comorbs broken into:
    # -   HTN T / F
    output_df["has_htn"] = database_df['Comorb'].apply(dz_is_in, args=("htn",))
    labels_ser["has_htn"] = "0 = no HTN, 1 = has HTN"

    # -   DM
    output_df["has_dm"] = database_df['Comorb'].apply(dz_is_in, args=("dm",))
    labels_ser["has_dm"] = "0 = no DM, 1 = has DM"

    # -   Psych
    output_df["has_psych"] = database_df['Comorb'].apply(dz_is_in, args=("psych",))
    labels_ser["has_psych"] = "0 = no psych, 1 = has psych"

    # -   Renal
    output_df["has_ckd"] = database_df['Comorb'].apply(dz_is_in, args=("ckd",))
    labels_ser["has_ckd"] = "0 = no CKD, 1 = has CKD"

    # -   Heart?
    #output_df["has_cv"] = database_df["Heart"].apply(is_dz_free)
    #labels_ser["has_cv"] = "0 = no CV disease, 1 = has some CV disease"

    output_df["has_cv"] = database_df['PostDx'].apply(dz_is_in, args=("Cardiac",))
    labels_ser["has_cv"] = "0 = no CV disease, 1 = has some CV disease"

    # -   CNS
    #output_df["has_cns"] = database_df["CNS"].apply(is_dz_free)
    #labels_ser["has_cns"] = "0 = no CNS disease, 1 = has some CNS disease"

    output_df["has_cns"] = database_df['PostDx'].apply(dz_is_in, args=("Neurologic",))
    labels_ser["has_cns"] = "0 = no CNS disease, 1 = has some CNS disease"

    # -   Opiate
    output_df["has_opiate"] = database_df['PostDx'].apply(dz_is_in, args=("Medication",))
    labels_ser["has_opiate"] = "0 = no opiate, 1 = on opiates"

    # AHI no coding
    output_df["AHI"] = database_df["AHI"]
    labels_ser["AHI"] = "no coding"

    # HFrEF y/n
    output_df["has_hfref"] = database_df["Heart"].apply(dz_is_in, args=("hfref",))
    labels_ser["has_hfref"] = "0 = no hfref, 1 = has hfref"

    #TODO: HfpEF y/n? Afib y/n?

    # HFpEF and AF
    output_df["has_hfpef_and_af"] = database_df["Heart"].apply(dzs_are_in, args=("hfpef", "afib",))
    labels_ser["has_hfpef_and_af"] = "0 = no hfpef or no afib, 1 = has both hfpef and af"

    # HFrEF and AF y/n
    output_df["has_hfref_and_af"] = database_df["Heart"].apply(dzs_are_in, args=("hfref", "afib",))
    labels_ser["has_hfref_and_af"] = "0 = no hfref or no afib, 1 = has both hfref and af"

    # Stroke y/n
    output_df["has_cva"] = database_df["CNS"].apply(dz_is_in, args=("cva",))
    labels_ser["has_cva"] = "0 = no cva, 1 = has had a cva"

    # Dementia or neurodegen

    # output_df["has_dementia"] = database_df["CNS"].apply(dz_is_in, args=("dementia",))
    # labels_ser["has_dementia"] = "0 = no dementia, 1 = has dementia"
    # output_df["has_neurodegen"] = database_df["CNS"].apply(dz_is_in, args=("neurodegenerative",))
    # labels_ser["has_neurodegen"] = "0 = no neurodegenerative disorder, 1 = has neurodegenerative disorder"

    # then add those two together, then round 2 (= has both) back to 1 (= has any)
    output_df["has_dem_or_neurodegen"] = (database_df["CNS"].apply(dz_is_in, args=("dementia",)) + \
                                         database_df["CNS"].apply(dz_is_in, args=("neurodegenerative",))).replace(2, 1)
    labels_ser["has_dem_or_neurodegen"] = "0 = no dementia or neurodegen, 1 = has dementia or neurodegen"

    # Dementia and (stroke or neurodegen)

    output_df["has_neurodegen_and_cva"] = database_df["CNS"].apply(dzs_are_in, args=("neurodegenerative", "cva",))
    labels_ser["has_neurodegen_and_cva"] = "0 = no cva or no neurodegen, 1 = has both cva and neurodegen"
    output_df["has_dem_and_cva"] = database_df["CNS"].apply(dzs_are_in, args=("dementia", "cva",))
    labels_ser["has_dem_and_cva"] = "0 = no cva or no dementia, 1 = has both cva and dementia"

    # then add those two together, then round 2 (= has both) back to 1 (= has any)
    output_df["has_dem_and_cva_or_degen"] = (database_df["CNS"].apply(dzs_are_in, args=("neurodegenerative", "cva",)) + \
                                            database_df["CNS"].apply(dzs_are_in, args=("dementia", "cva",))).replace(2,1)
    labels_ser["has_dem_and_cva_or_degen"] = "1 = has (dem and cva) or (neurodegen and cva), 0 = doesn't have those combos"

    # Final Treatment

    # #Similarly do the same for ASV (group 0) vs. CPAP and BPAP together.
    database_df['FinalTx_coll'] = database_df.apply(collapse_final_treatment, axis=1)

    LE.fit(database_df['FinalTx_coll'])
    class_for_swap = list(LE.classes_)  # swapped the order of the labels to be more intuitive
    class_for_swap[0], class_for_swap[1] = class_for_swap[1], class_for_swap[0]
    labels_ser['FinalTx_coll'] = create_key_string(class_for_swap)  #  switch labels for PAP to be 0 (not 1) and ASV to be 1 (not 0)

    output_df['FinalTx_coll'] = LE.transform(database_df['FinalTx_coll'])
    output_df['FinalTx_coll'] = output_df['FinalTx_coll'].apply(swap_value)  #so that the labels are correct order

    # Collapse: percOSA/CSA as 0 and 1 (0 being >50% OSA –combine 0 and 1 groups and 1 being >50% CSA –combine the 3 and 4).

    database_df['PercOSA'] = database_df.apply(collapse_base_dx, axis=1)

    # Category of OSA
    LE.fit(database_df['PercOSA'])
    class_for_swap = list(LE.classes_)  # swapped the order of the labels to be more intuitive
    labels_ser['PercOSA'] = create_key_string(class_for_swap[1:] + [class_for_swap[0]])  #  Perc(entage) OSA: more descriptive term for BaseDx
    output_df['PercOSA'] = LE.transform(database_df['PercOSA'])
    output_df['PercOSA'] = output_df['PercOSA'].apply(swap_value)  #so that the labels are correct order

    # Dx Study
    LE.fit(database_df['StudyType'])
    labels_ser['StudyType'] = create_key_string(list(LE.classes_))
    output_df['StudyType'] = LE.transform(database_df['StudyType'])

    output_df.to_excel(output_loc)
    labels_ser.to_excel("keys_"+output_loc)
    return


def swap_value(value):
    '''make 0 a 1 and a 1 a 0 for when SKLearn gives the backward encoding
    if value is anything other than 1 or 0, no swap is performed'''
    if value == 1:
        value = 0
    elif value == 0:
        value = 1
    return value


def dz_is_in(dz_string, substring):
    """ returns 0 if substring is NOT in dz_string
    returns 1 if substring is in dz_string"""
    if substring not in dz_string:
        return 0
    else:
        return 1


def dzs_are_in(dz_string, substring1, substring2):
    """ returns 1 if both substrings are in dz_string
    otherwise, returns 0"""
    if substring1 not in dz_string:
        return 0
    elif substring2 not in dz_string:
        return 0
    else:
        return 1


def is_dz_free(dz_string, substring="none"):
    """ returns 0 if the return substring is in dz_string
    default substring is none, meaning that if CV = none, patient is disease free
    returns 1 if substring is not in dz_string"""
    if substring not in dz_string:
        return 1
    else:
        return 0


def replace_etiology_labels(df):
    """take a df of histogram (e.g. count_hist) and replace the labels such that:
     cardiac -> cardiac conditions
     neurologic -> neurologic conditions
     medication -> opiate use
     Primary -> primary CSA"""
    return df.replace({'Cardiac':'Cardiac Conditions', 'Neurologic':'Neurologic Conditions', 'Medication':'Opiate Use', 'Primary':'Primary CSA'})


def collapse_initial_outcome(row):
    # take a row and returns either "Failed CPAP", "Resolved with CPAP", "No Adequate CPAP Trial"
    if row['Outcome'] == "n/a":
        return "No adequate CPAP trial"
    elif row["Outcome"] == "never started on cpap":
        return "No adequate CPAP trial"
    elif row["Outcome"] == "non-compliant":
        return "No adequate CPAP trial"
    elif row["Outcome"] == "resolved w/bipap":
        return "No adequate CPAP trial"
    elif row["Outcome"] == "resolved w/ cpap":
        return "Resolved with CPAP"
    elif row["Outcome"] == "failed cpap":
        return "Did not resolve with CPAP"
    else:
        return row["Outcome"]


def collapse_final_treatment(row):
    #takes a row and returns either
    # CPAP, BPAP, ASV, or other
    #print(row['FinalTx'])
    if row['FinalTx'] == 'niv-o2':
        return "ASV"
    elif row['FinalTx'] == 'niv':
        return "ASV"
    elif row['FinalTx'] == 'asv':
        return "ASV"
    elif row['FinalTx'] == 'bipap-o2':
        return "BPAP"
    elif row['FinalTx'] == 'bipap':
        return "BPAP"
    elif row['FinalTx'] == 'cpap':
        return "CPAP"
    elif row['FinalTx'] == 'O2':
        return "Other"
    elif row['FinalTx'] == 'mad':
        return "Other"
    elif row['FinalTx'] == 'none':
        return "Other"
    else:
        print("ERROR, collapse_final_treatment")
        return row['FinalTx']


def collapse_base_dx(row):
    #takes a row and returns either
    # mostly_CSA or mostly_OSA depending on if BaseDx has more or less than 50% obstructive/central apneas
    #print(row['BaseDx'])
    if row['BaseDx'] == 'Mainly OSA':
        return "mostly_OSA"
    elif row['BaseDx'] == 'Combined OSA/CSA':
        return "mostly_OSA"
    elif row['BaseDx'] == 'Predominantly CSA':
        return "mostly_CSA"
    elif row['BaseDx'] == 'Pure CSA':
        return "mostly_CSA"
    else:
        print("ERROR, collapse_base_dx")
        return row['BaseDx']


def create_key_string(classes):
    '''takes the list of classes from the LabelEncoder and puts them in to a string to function as a key,
     and subsequently be appended to the output encoded dataframe at the end'''

    output_str = ""
    for i in range(len(classes)):
        output_str += str(i) + " = " + str(classes[i]) + ",  "
    return output_str


def display_dist(df, label):
    """takes a df and a column label and graphs the distribution (continuous) for display"""
    # print(sns.__version__)

    #TODO: make a categorical version of this?

    sns.set(style="white", palette="pastel")
    fig, axes = plt.subplots(2, 1, figsize=(6, 6))  # 6, 2 if OSA CSA

    axes[1].get_shared_x_axes().join(axes[1], axes[0])
    axes[1].set_aspect(aspect=25)

    sns.distplot(df[label], ax=axes[0], kde=False, norm_hist=False, color='teal')
    sns.boxplot(data=df, x=label, ax=axes[1], color='skyblue')

    sns.despine(ax=axes[0], top=True, bottom=True, right=True)
    sns.despine(ax=axes[1], top=True, left=True, right=True)

    axes[0].set_xlabel("")
    axes[0].set_ylabel("Count per bin", fontsize='large')

    row_label = "{lab}\nMean: {mean:.1f}, Std Dev: {std:.1f}\nMedian: {med:.1f}, IQR: [{lower:.1f}, {upper:.1f}]\nCount: {count:.0f}"\
        .format(lab=label, mean=df[label].describe()['mean'], std=df[label].describe()['std'],
                med=df[label].describe()['50%'], lower=df[label].describe()['25%'], upper=df[label].describe()['75%'],
                count=df[label].describe()['count'])

    axes[1].set_xlabel(row_label, fontsize='large')
    axes[1].set(xlim=(0, None))

    fig.suptitle("Distribution of " + str(label), fontsize='xx-large')
    fig.tight_layout(rect=[0, 0, 1, .9])  # .95 to leave space for title
    fig.savefig('Display Dist ' + str(label) + '.png', dpi=100)


def figure_4(df):
    """
    Bar graphs of the final treatment by percentage of CSA
    """

    sns.set(style="white", palette=sns.color_palette("cubehelix", 6))
    f, axes = plt.subplots(4, 1, figsize=(6, 9))  # , sharex=True)
    sns.despine(top=True, bottom=True)

    # Add collapsed final treatments
    df['FinalTx_coll'] = df.apply(collapse_final_treatment, axis=1)

    neurologic_df = df.loc[df['PostDx'].str.contains("Neurologic")].sort_values(by='Outcome')

    OSA_pure_hist = df.loc[df['BaseDx'] == "Mainly OSA"].sort_values("FinalTx_coll")
    OSA_predom_hist = df.loc[df['BaseDx'] == "Combined OSA/CSA"].sort_values("FinalTx_coll")
    CSA_predom_hist = df.loc[df['BaseDx'] == "Predominantly CSA"].sort_values("FinalTx_coll")
    CSA_pure_hist = df.loc[df['BaseDx'] == "Pure CSA"].sort_values("FinalTx_coll")

    # Create count plot for each #CSA on the left, then a Pie Chart with proportion on the right

    hatches = ['', '', '\\\\\\', '...']
    face_color = ['dimgrey','white', 'white', 'white']

    # Pure OSA
    bar = sns.countplot(y="FinalTx_coll", data=OSA_pure_hist, ax=axes[3])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[3].set(xlabel="", ylabel="<10% CSA")

    # Predom OSA
    bar = sns.countplot(y="FinalTx_coll", data=OSA_predom_hist, ax=axes[2])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[2].set(xlabel="", ylabel="10-49.9% CSA")

    # Predom CSA
    bar = sns.countplot(y="FinalTx_coll", data=CSA_predom_hist, ax=axes[1])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[1].set(xlabel="", ylabel="50-90% CSA")

    # Pure CSA
    bar = sns.countplot(y="FinalTx_coll", data=CSA_pure_hist, ax=axes[0])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[0].set(xlabel="", ylabel=">90% CSA")

    # Combined X axis for L side
    axes[3].get_shared_x_axes().join(axes[3], axes[2], axes[1], axes[0])
    axes[0].set_xticklabels("")
    axes[1].set_xticklabels("")
    axes[2].set_xticklabels("")
    # Leave bottom aka [3,0] labels in

    # Resize all
    axes[0].autoscale()
    axes[1].autoscale()
    axes[2].autoscale()
    axes[3].autoscale()

    f.tight_layout(rect=[0, 0, 1, 1])  # .95 to leave space for title
    f.savefig('Figure 4 - final tx by perc csa', dpi=100)
    # plt.show()


def figure_3(df):
    """
    Bar graphs of the outcome of CPAP treatment (if performed) by the etiology of CSA
    """

    sns.set(style="white", palette=sns.color_palette("cubehelix", 6)) #still 6 if osa csa
    f, axes = plt.subplots(5, 1, figsize=(6, 9)) # 6, 2 if OSA CSA
    sns.despine(top=True, bottom=True)
    # f.suptitle("Outcome, Grouped by Contributing Etiology")

    # contains used instead of equal to include patients with multiple etio (e.g. cardiac+medication count to both)
    neurologic_df = df.loc[df['PostDx'].str.contains("Neurologic")].sort_values(by='Outcome')
    cardiac_df = df.loc[df['PostDx'].str.contains("Cardiac")].sort_values(by='Outcome')
    medication_df = df.loc[df['PostDx'].str.contains("Medication")].sort_values(by='Outcome')
    tecsa_df = df.loc[df['PostDx'].str.contains("TECSA")].sort_values(by='Outcome')
    # osacsa_df = df.loc[df['PostDx'].str.contains("OSA-CSA")].sort_values(by='Outcome')
    primary_df = df.loc[df['PostDx'].str.contains("Primary")].sort_values(by='Outcome')

    # collapse possible outcomes
    neurologic_df['col_outcome'] = neurologic_df.apply(collapse_initial_outcome, axis=1)
    cardiac_df['col_outcome'] = cardiac_df.apply(collapse_initial_outcome, axis=1)
    medication_df['col_outcome'] = medication_df.apply(collapse_initial_outcome, axis=1)
    tecsa_df['col_outcome'] = tecsa_df.apply(collapse_initial_outcome, axis=1)
    # osacsa_df['col_outcome'] = osacsa_df.apply(collapse_initial_outcome, axis=1)
    primary_df['col_outcome'] = primary_df.apply(collapse_initial_outcome, axis=1)

    # Create count plot for each Etio on the left, then a Pie Chart with proportion on the right

    hatches = ['', '||||', '']
    face_color = ['white', 'white', 'dimgrey']

    # Neurologic
    bar = sns.countplot(y='col_outcome', data=neurologic_df, ax=axes[0])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[0].set(xlabel="", ylabel="Neurologic\nConditions")


    # Cardiac
    bar = sns.countplot(y='col_outcome', data=cardiac_df, ax=axes[1])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[1].set(xlabel="", ylabel="Cardiac\nConditions")

    # Medication
    bar = sns.countplot(y='col_outcome', data=medication_df, ax=axes[2])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[2].set(xlabel="", ylabel="Opiate Use")

    # OSA-CSA
    # bar = sns.countplot(y='col_outcome', data=osacsa_df, ax=axes[3,0])
    # for i, this_bar in enumerate(bar.patches):
    #     # Set a different hatch for each bar
    #     this_bar.set_hatch(hatches[i])
    # axes[3].set(xlabel="", ylabel="OSA-CSA")
    # If adding OSA-CSA back, would need to increase by 1 all of the axes indices

    # TE-CSA
    bar = sns.countplot(y='col_outcome', data=tecsa_df, ax=axes[3])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[3].set(xlabel="", ylabel="TECSA")

    #Primary
    bar = sns.countplot(y='col_outcome', data=primary_df, ax=axes[4])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[4].set(xlabel="", ylabel="Primary CSA")

    # Combined X axis for L side
    axes[4].get_shared_x_axes().join(axes[4], axes[3], axes[2], axes[1], axes[0]) # axes[5] would need to be added back
    axes[0].set_xticklabels("")
    axes[1].set_xticklabels("")
    axes[2].set_xticklabels("")
    axes[3].set_xticklabels("")
    # axes[4].set_xticklabels("")
    # Leave bottom labels in

    # Resize all
    axes[0].autoscale()
    axes[1].autoscale()
    axes[2].autoscale()
    axes[3].autoscale()
    axes[4].autoscale()
    # axes[5].autoscale()

    f.tight_layout(rect=[0, 0, 1, 1])
    f.savefig('Figure 3 - outcome of cpap by etio.png', dpi=100)
    # plt.show()

def figure_2(df):
    """
    bar graphs of etiology (e.g. cardiac, medication) subdivided by the percentage of CSA
    """

    sns.set(style="white", palette=sns.color_palette("cubehelix", 6))
    f, axes = plt.subplots(4, 1, figsize=(6, 9))  # , sharex=True)
    sns.despine(top=True, bottom=True)
    # f.suptitle("Etiology of Central Events, Grouped by %Central Events")

    OSA_pure_hist = replace_etiology_labels(histo_dx_includes(df.loc[df['BaseDx'] == "Mainly OSA"], return_df=True).sort_values("Dx"))
    OSA_predom_hist = replace_etiology_labels(histo_dx_includes(df.loc[df['BaseDx'] == "Combined OSA/CSA"], return_df=True).sort_values("Dx"))
    CSA_predom_hist = replace_etiology_labels(histo_dx_includes(df.loc[df['BaseDx'] == "Predominantly CSA"], return_df=True).sort_values("Dx"))
    CSA_pure_hist = replace_etiology_labels(histo_dx_includes(df.loc[df['BaseDx'] == "Pure CSA"], return_df=True).sort_values("Dx"))

    # Create count plot for each #CSA on the left, then a Pie Chart with proportion on the right

    hatches = ['///', 'xxx', '---', '', '']
    face_color = ['white', 'white', 'white', 'white', 'dimgrey']

    # Pure OSA
    bar = sns.barplot(x="Count", y="Dx", data=OSA_pure_hist, ax=axes[3])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[3].set(xlabel="", ylabel="<10% CSA")

    # Predom OSA
    bar = sns.barplot(x="Count", y="Dx", data=OSA_predom_hist, ax=axes[2])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[2].set(xlabel="", ylabel="10-49.9% CSA")

    # Predom CSA
    bar = sns.barplot(x="Count", y="Dx", data=CSA_predom_hist, ax=axes[1])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[1].set(xlabel="", ylabel="50-90% CSA")

    # Pure CSA
    bar = sns.barplot(x="Count", y="Dx", data=CSA_pure_hist, ax=axes[0])
    for i, this_bar in enumerate(bar.patches):
        # Set a different hatch for each bar
        this_bar.set_edgecolor('black')
        this_bar.set_facecolor(face_color[i])
        this_bar.set_hatch(hatches[i])
    axes[0].set(xlabel="", ylabel=">90% CSA")

    # Combined X axis for L side
    axes[3].get_shared_x_axes().join(axes[3], axes[2], axes[1], axes[0])
    axes[0].set_xticklabels("")
    axes[1].set_xticklabels("")
    axes[2].set_xticklabels("")
    # Leave bottom aka [3,0] labels in

    # Resize all
    axes[0].autoscale()
    axes[1].autoscale()
    axes[2].autoscale()
    axes[3].autoscale()

    f.tight_layout(rect=[0, 0, 1, 1])  # .95 to leave space for title
    f.savefig('Figure 2 - etio by perc csa', dpi=100)
    # plt.show()


def main():
    # Location of Db file
    db_loc = "/Users/reblocke/Box/Residency Personal Files/Scholarly Work/CSA/Databases/CSA-Db-Working.xlsm"
    # db_loc = "/Users/reblocke/Box/Residency Personal Files/Scholarly Work/CSA/Databases/Backups/CSA-Db-Working-Only OSA CSA no inpt.xlsm"
    # uncomment row 377 of readexcel.py if using above (includes OSA-CSA)
    df = arrays_to_df(sheet_to_arrays(load_sheet(db_loc)))

    df.to_excel('output.xlsx')

    coded_output(df)
    new_make_tables(df)

    figure_2(df)
    figure_3(df)
    figure_4(df)

    # print("\n\n---Total of number of patients where each etiology was contributory---")
    # print("---(will some to more than total given mutliple dx's)---\n")

    # Visualization tester function
    # visualizations(df)

    # Other visualizations:
    # vis_hist_etio(df)
    # init_tx_by_csa(df)
    # final_tx_by_csa(df)
    # init_tx_by_etio(df)
    # outcome_by_etio(df)
    # outcome_by_etio_no_pie(df)
    # etio_by_csa(df)
    # etio_by_csa_no_pie(df)
    # pieChartBaseDx(df)
    # sankeyEtioTx(df)
    # sankeyTypeFinalTx(df)
    # sankeyTypeOutcome(df)
    # outcome_by_csa_percent(df)
    # sankeyTypeFinalTx(df)
    # test_by_etio(df)
    # test_by_csa(df)

    display_dist(df, 'Age')

if __name__ == '__main__':
    main()
