# Read Excel Files for the CSA project
import openpyxl as op
import pandas as pd

def load_sheet(location):
    """loads the sheet excel doc and returns the 1st sheet"""
    wb = op.load_workbook(location, read_only=True, data_only=True)
    return wb[wb.sheetnames[0]]

def sheet_to_arrays(excel_sheet):
    """makes a 2 dimensional list database and fills it based on an
    openpyxl excel sheet"""
    # Name_Column = 1
    # MRN_Column = 2
    # DOB_Column = 3
    Age_Column = 4  # Age at diag sleep study
    Sex_Column = 5
    Race_Column = 6
    # Zip_Column = 7
    BMI_Column = 8
    Smoking_Column = 9
    Comorb_Column = 10
    Heart_Column = 11
    CNS_Column = 12
    Base_Dx_Column = 13
    AHI_Column = 14  # Diagnostic AHI
    Post_Dx_Column = 15
    Final_Tx_Column = 16
    Outcome_Column = 17
    Path_ASV_Column = 18
    Time_ASV_Column = 19
    # Loc_Column = 20
    Sleep_Study_Column = 21

    Patients = list()

    # print("Processing excel spreadsheet")
    i = 1

    for patient in excel_sheet.iter_rows(max_row=513):  # not sure why I needed to manually hardcode this - size of db
        # For each row that has an MRN entry...
        # print("Processing chart #" + str(i))
        row = list()
        row.append(i-1) # = ID, previously (patient[MRN_Column].value)

        try:
            row.append(int(patient[Age_Column].value))
        except(ValueError, TypeError, AttributeError):
            print("Age Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(patient[Sex_Column].value.lower()) #Make discrete?
        except(ValueError, TypeError, AttributeError):
            print("Sex Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(patient[Race_Column].value.lower()) #Make discrete?
        except(ValueError, TypeError, AttributeError):
            print("Race Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(patient[Smoking_Column].value.lower())  # Make discrete?
        except(ValueError, TypeError, AttributeError):
            print("Smoking Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(float(patient[BMI_Column].value))
        except(ValueError, TypeError, AttributeError):
            print("BMI Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(patient[Comorb_Column].value.lower().strip())
        except(ValueError, TypeError, AttributeError):
            print("Comorbidity Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(patient[Heart_Column].value.lower().strip())
        except(ValueError, TypeError, AttributeError):
            print("Heart Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(patient[CNS_Column].value.lower().strip())
        except(ValueError, TypeError, AttributeError):
            print("CNS Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(float(patient[AHI_Column].value))
        except(ValueError, TypeError, AttributeError):
            print("AHI Column Error: Row " + str(i))
            row.append(None)
        # Should these following ones be made discrete?
        try:
            row.append(patient[Base_Dx_Column].value.lower().strip())
        except(ValueError, TypeError, AttributeError):
            print("Base Dx Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(patient[Post_Dx_Column].value.lower().strip())
        except(ValueError, TypeError, AttributeError):
            print("Post DX Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(patient[Final_Tx_Column].value.lower().strip())
        except(ValueError, TypeError, AttributeError):
            print("Final Tx Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(patient[Outcome_Column].value.lower().strip())
        except(ValueError, TypeError, AttributeError):
            print("Outcome Column Error: Row " + str(i))
            row.append(None)
        try:
            row.append(patient[Path_ASV_Column].value.lower().strip())
        except(ValueError, TypeError, AttributeError):
            print("Path to ASV Column Error: Row " + str(i))
            row.append(None)

        try:
            row.append(patient[Time_ASV_Column].value.lower().strip())
        except(ValueError, TypeError, AttributeError):
            print("Time to ASV Column Error: Row " + str(i))
            row.append(None)

        try:
            row.append(patient[Sleep_Study_Column].value.lower().strip())
        except(ValueError, TypeError, AttributeError):
            print("Sleep Study Type Column Error: Row " + str(i))
            row.append(None)
        Patients.append(row)
        i = i+1

    # print(Patients[1:])
    return Patients[1:]  # take off the first row = labels


def histo_dx_includes(df, return_df=False):
    """Returns a histogram (pandas series) of diagnosis where a post-titration diagnosis of
    w/ multiple factors (e.g. Meds+CV) each are counted toward their respective
    category counts

    If return_df = true, will return in dataframe (with 'Dx', 'Count' as keys) instead of a series"""

    histo = pd.Series({"TECSA":0,
        # "OSA-CSA":0, Removed for now
        "Cardiac":0,
        "Neurologic":0,
        "Medication":0,
        "Primary":0})

    for dx in df['PostDx']:
        dxstr = str(dx)
        for cat in histo.index:
            if cat in dxstr:
                histo[cat] +=1
    histo = histo.sort_values(ascending=False)
    if return_df is False:
        return histo
    else:
        return pd.DataFrame({"Dx": histo.index, "Count": histo.values})

def histo_comorbs_includes(df):
    """Returns a histogram (pandas series) of comorbidities where a comorbidity of
    w/ multiple factors (e.g. CKD+Psych) each are counted toward their respective
    category counts"""
    histo = pd.Series({"none":0,
        "htn":0,
        "dm":0,
        "ckd":0,
        # "hiv":0,
        "psych":0})

    for comorb in df['Comorb']:
        comorb_str = str(comorb)
        for cat in histo.index:
            if cat in comorb_str:
                histo[cat] +=1
    return histo.sort_values(ascending=False)

def histo_heart_includes(df):
    """Returns a histogram (pandas series) of heart comorbidities where a comorbidity of
    w/ multiple factors (e.g. afib and cad) each are counted toward their respective
    category counts"""
    histo = pd.Series({"none":0,
        "cad":0,
        "afib":0,
        "hfpef":0,
        "hfref":0,
        "other":0})

    for heart in df['Heart']:
        heart_str = str(heart)
        for cat in histo.index:
            if cat in heart_str:
                histo[cat] +=1
    return histo.sort_values(ascending=False)


def histo_cns_includes(df):
    """Returns a histogram (pandas series) of cns comorbidities where a comorbidity of
    w/ multiple factors (e.g. cva and dementia) each are counted toward their respective
    category counts"""
    histo = pd.Series({"none":0,
        "cva":0,
        "neurodegenerative":0,
        "dementia":0,
        # "seizures":0,
        # "mass":0,
        "chiari":0,
        "other":0})

    for cns in df['CNS']:
        cns_str = str(cns)
        for cat in histo.index:
            if cat in cns_str:
                histo[cat] +=1
    return histo.sort_values(ascending=False)


def arrays_to_df(patient_array):
    """takes the database in array form and outputs a dataframe with variables
    categorized.

    ['ID', Age', 'Sex', 'BMI', 'AHI', 'BaseDx', 'PostDx', 'FinalTx', 'Outcome',
    "ProcToASV", "TimeToASV"]

    Also adds an inferred column "InitialTx" """

    df = pd.DataFrame.from_records(patient_array, columns=['ID', 'Age',  'Sex', 'Race', 'Smoking', 'BMI', 'Comorb',
                                                           'Heart', 'CNS', 'AHI', 'BaseDx', 'PostDx', 'FinalTx',
                                                           'Outcome', "ProcToASV", "TimeToASV", "StudyType"])  #AHI_label added after

    df['Sex'] = df['Sex'].astype('category')

    df['Race'] = df['Race'].astype('category')
    df['Race'] = df['Race'].replace({"not hispanic/ latino": 'not hispanic/latino'})
    df['Race'] = df['Race'].replace({"not hispanic/latino": 'white'})  # Utah Adjustment :(

    df['Smoking'] = df['Smoking'].astype('category')

    df['Comorb'] = df['Comorb'].apply(matchComorbs).astype('category')

    df['Heart'] = df['Heart'].apply(matchHeart).astype('category')

    df['CNS'] = df['CNS'].apply(matchCNS).astype('category')

    AHILabelCat = pd.api.types.CategoricalDtype(categories=['none', 'mild', 'moderate', 'severe'], ordered=True)
    df['AHI_label'] = df['AHI'].apply(ahi_label).astype(AHILabelCat)

    df['BaseDx'] = df['BaseDx'].replace(
        {"Mainly OSA (<10% CSA or most centra events either SOCAPACA)".lower(): 'Mainly OSA',
        "Combined OSA/CSA (CSA 10-50%)".lower(): 'Combined OSA/CSA',
        "Predominantly CSA (>50% CSA)".lower(): 'Predominantly CSA',
        "Pure CSA (<10% OSA)".lower(): 'Pure CSA'})
    BaseDxCat = pd.api.types.CategoricalDtype(categories=[
        'Mainly OSA', 'Combined OSA/CSA', 'Predominantly CSA',
        'Pure CSA'], ordered=True)
    df['BaseDx'] = df['BaseDx'].astype(BaseDxCat)

    # transform PostDx to shorter labels
    df['PostDx'] = df['PostDx'].apply(matchDx).astype('category')

    df['FinalTx'] = df['FinalTx'].replace(
        {"cpap": "cpap",
        "bipap": "bipap",
        "bipap w/ o2": "bipap-o2",
        "asv (resmed/ respironics)": "asv",
        "supplemental oxygen": "O2",
        "no treatment": "none",
        "MAD":"mad",
        "avaps": "niv",
        "ivaps w/ o2": "niv-o2",
        "ivaps": "niv"})
    FinalTxCat = pd.api.types.CategoricalDtype(categories=["niv-o2", "niv", "asv",
        "bipap-o2", "bipap", "cpap", "mad", "O2", "none"], ordered=True)
    df['FinalTx'] = df['FinalTx'].astype(FinalTxCat)
#    OutcomeCat = pd.api.types.CategoricalDtype(categories=[
#        "resolved w/ cpap", "failed cpap", "non-compliant", "n/a"], ordered=False)

    df['Outcome'] = df['Outcome'].astype('category')

    df['ProcToASV'] = df['ProcToASV'].replace(
        {"n/a": 'other',
        "initial treatment": "initial treatment",
        "after trial of cpap": "after trial of cpap",
        "after trial of bipap": 'after trial of bipap'
        })
    procToASVCat = pd.api.types.CategoricalDtype(categories=['other',
        "initial treatment", 'after trial of cpap', 'after trial of bipap'],
        ordered=True)
    df['ProcToASV'] = df['ProcToASV'].astype(procToASVCat)

    df['TimeToASV'] = df['TimeToASV'].replace(
        {"n/a": 'other',
        "0-1 month": 'within 2 mo',
        "3-6 months": '3-6 mo',
        ">6 months": '6+ mo'})
    timeToASVCat = pd.api.types.CategoricalDtype(categories=['other',
        "within 2 mo", '3-6 mo', '6+ mo'], ordered=True)
    df['TimeToASV'] = df['TimeToASV'].astype(timeToASVCat)

    df['InitTx'] = df.apply(infer_initial_treatment, axis=1)
    initTxCat = pd.api.types.CategoricalDtype(categories=["asv", "cpap", "not cpap/asv (or unknown)"], ordered=True)
    df['InitTx'] = df['InitTx'].astype(initTxCat)

    df['StudyType'] = df['StudyType'].astype('category')



    return df


def blow_out_comorbs_cv_neuro(df):
    '''takes the finished (preprocessed) database dataframe and separates all the comorbidity, heart disease, and cns
    disease into individual disorders and true vs false whether the patient has them or not

    e.g. instead of Psych+HTN would be
    psych true
    HTN true
    DM false
    etc.

    Adds these as separate columns

        histo = pd.Series({"none":0,
        "htn":0,
        "dm":0,
        "ckd":0,
        # "hiv":0,
        "psych":0})

        {"none":0,
        "cad":0,
        "afib":0,
        "hfpef":0,
        "hfref":0,
        "other":0}

        {"none":0,
        "cva":0,
        "neurodegenerative":0,
        "dementia":0,
        # "seizures":0,
        # "mass":0,
        "chiari":0,
        "other":0}
        '''

    pass


def infer_initial_treatment(patient):
    """takes a patient (row) from the dataframe and infers what the initial treatment was, returned as string"""
    init_tx = 'not cpap/asv (or unknown)'    # default
    if patient['FinalTx'] == "cpap":
        init_tx = "cpap"
    if patient['FinalTx'] == "bipap" or patient['FinalTx' == "bipap-o2"]:
        if patient['Outcome'] == 'failed cpap':
            init_tx = "cpap"
    if patient['FinalTx'] == "asv":
        if patient['ProcToASV'] == "initial treatment":
            init_tx = "asv"
        elif patient['ProcToASV'] == 'after trial of cpap':
            init_tx = "cpap"    # Note: this is an assumption: all patients who trialed CPAP prior to ASV started w CPAP
        elif patient['Outcome'] == "failed cpap":
            init_tx = "cpap"    # Note: this is an assumption: all patients who failed CPAP prior to ASV started w CPAP
    if patient['FinalTx'] == "none" or patient['FinalTx'] == "O2" or patient['FinalTx'] == "other":
        if patient['Outcome'] == 'failed cpap' or patient['Outcome'] == "never started cpap" \
                or patient['Outcome'] == "resolved w/ cpap":
            init_tx = "cpap"
    return init_tx


def matchDx(pt_dx):
    """match the diagnosis up with the shorter labels"""
    new_dx = list()
    rep = {"te csa": "TECSA",
        "csa w/cns dz (tbi/ cerebrovascular dz/ mass lesion/ neurodegenerative dz/ other)":"Neurologic",
        "primary csa (idiopathic csa)":"Primary",
        #"osa-associated":"OSA-CSA",  #  removed, as these excluded now
        "csa w/opioid (methadone/ fentanyl/ oxycontin/ suboxone/ other)":"Medication",
        "csa w/heart dz (hfref <45%/ hfpef >45% /a.fib)":"Cardiac"}
    for dx in pt_dx.split(","):
        new_dx.append(rep[dx.strip().lower()])  # transform labels
    return '+'.join(sorted(new_dx))   # make sure that order doesn't matter, join iterable list


def ahi_label(ahi):
    """return the label for the severity of OSA based on AHI"""
    if ahi <= 5.0:
        return "mild"  # previously "none"
    elif ahi < 15.0:
        return "mild"
    elif ahi < 30.0:
        return "moderate"
    elif ahi >= 30.0:
        return "severe"
    else:
        return "mild"  #  previously "error"  # shouldn't happen, will cause flag at conversion to type


def matchComorbs(pt_comorb):
    """match the comorbidities up with the shorter labels"""
    new_comorb = list()
    rep = {"htn": "htn",
           # "hiv": "hiv",
           "dm": "dm",
           "psychiatric": "psych",
           "renal failure (creatinine>2mg/dl/ use of rrt/ cr clearance <30ml/min": "ckd",
           "none": "none"}
    for comorb in pt_comorb.split(","):
        new_comorb.append(rep[comorb.strip().lower()])
    return '+'.join(sorted(new_comorb))


def matchHeart(pt_heart):
    """match the heart comorbidities up with the shorter labels"""
    new_heart = list()
    rep = {"cad": "cad",
           "atrial fibrillation": "afib",
           "chf- hfpef (>45%)": "hfpef",
           "chf- hfref (<45%)": "hfref",
           "pacs": "other",
           "svt": "other",
           "atrial myxoma": "other",
           "cardiac tx": "other",
           "avnrt": "other",
           "none": "none"}
    for heart in pt_heart.split(","):
        new_heart.append(rep[heart.strip().lower()])
    return '+'.join(sorted(new_heart))


def matchCNS(pt_cns):
    """match the heart comorbidities up with the shorter labels"""
    new_cns = list()
    rep = {"ischemic stroke": "cva",
           "neurodegenerative disease": "neurodegenerative",
           "tbi": "other",
           "dementia": "dementia",
           "seizure disorder": "other",
           "mass lesion": "other",
           "tia": "cva",
           "chiari malformation": "chiari",
           "traumatic brain injury": "tbi",
           "ms": "neurodegenerative",
           "epilepsy": "seizures",
           "pituitary adenoma": "other",
           "hemorrhagic stroke": "cva",
           "tumor": "other",
           "other": "other",
           "cerebral palsy": "other",
           "seizures": "seizures",
           "none": "none"}
    for cns in pt_cns.split(","):
        new_cns.append(rep[cns.strip().lower()])
    return '+'.join(sorted(new_cns))


def test_db_gen():
    # db = RecordsDb()
    pass


def main():
    # 0 for testing, 1 to run
    testing_mode = 1

    if testing_mode == 0:
        test_db_gen()
    else:
        # run the main program
        pass


if __name__ == '__main__':
    main()
